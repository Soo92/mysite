from io import BytesIO
from PIL import Image
import csv
import calendar
import datetime
from datetime import date, timedelta
import dateutil.relativedelta
import openpyxl
import clipboard
import requests
import hashlib, hmac, base64, requests, time, os
import urllib.request, json
from urllib.request import urlopen, Request
import random
import shutil
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.styles import Border, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import PIL
import io
import urllib3
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import zipfile
import os, time, numpy as np, pandas as pd
import xlsxwriter
from bs4 import BeautifulSoup    #BeautifulSoup import
import time

def set_df_to_sht(df,sht,y,x):
    dfy=y
    dfx=x
    for col in df.columns:
        sht.cell(dfy,dfx).value=col
        dfx=dfx+1
    for dfy in range(0,len(df)):
        dfx=x
        for col in df.columns:
            tmp_data=df.iloc[dfy][col]
            if col=="월간" or tmp_data=="-":
                tmp_data=str(tmp_data)
            elif "%" in str(tmp_data):
                tmp_data=float(tmp_data.replace("%",""))/100
            else:
                tmp_data=int(tmp_data)
            sht.cell(dfy+y+1,dfx).value=tmp_data
            dfx=dfx+1
    return dfy+y+1

def make_chart(ws,x_l,x_r,y1,y2,to_cell,h,w):
    cats = Reference(ws, min_col=2, max_col=2, min_row=y1, max_row=y2)
    values_l = Reference(ws, min_col=x_l, max_col=x_l, min_row=y1, max_row=y2)
    values_r = Reference(ws, min_col=x_r, max_col=x_r, min_row=y1, max_row=y2)

    chart_ad = LineChart()
    chart_ad.add_data(values_l)
    chart_ad.set_categories(cats)
    chart_ad.y_axis.axId = 200
    chart_ad.y_axis.title = ws.cell(14,x_l).value
    chart_ad.y_axis.majorGridlines = None
    s = chart_ad.series[0]
    # lineProp = drawing.line.LineProperties(prstDash='dash')
    s.graphicalProperties.line.prstDash = 'dash'
    s.graphicalProperties.line.solidFill = str(ws.cell(14,x_l).fill.start_color.index)[2:]

    chart_val = LineChart()
    chart_val.add_data(values_r)
    chart_val.y_axis.title = ws.cell(14,x_r).value

    chart_val.title = ws.cell(14,x_r).value
    chart_val.width = w
    chart_val.height = h
    chart_val.y_axis.crosses = "max"
    s = chart_val.series[0]
    s.graphicalProperties.line.solidFill = str(ws.cell(14,x_r).fill.start_color.index)[2:]
    chart_val += chart_ad
    chart_val.legend=None

    ws.add_chart(chart_val, to_cell)

def set_key_report():
# 엑셀데이터 가져오기
    ws_tmp=ws_key
    in_keyword_per_cut=df_admin_dict["키워드 상위(%)"]
    in_keyword_cnt_cut=df_admin_dict["키워드 최대갯수"]
    ws_tmp["C5"].value=in_keyword_per_cut
    ws_tmp["C6"].value=in_keyword_cnt_cut
# 인기키워드 top 25개중 10개 추출
    arr_key_rank=[20]
    arr_key_cnt=[10]
# 유입 csv 데이터 추출
    df=df_web_dict["검색어-보고서"][["검색어","노출수","클릭수","전환수"]]
    df=df[1:]
    columnNames = df.columns
    st_x=3
    st_y=7-1
    for i in range(1,4):
        col=columnNames[i]
        df[col] = df[col].astype(str).str.replace(',', '')
        df=df.astype({col: int})
        df=df.sort_values(by=col, ascending=False)
        df2=df[df[col] > df[col].quantile((100-in_keyword_per_cut)/100)]
        df2=df2.head(int(in_keyword_cnt_cut))
        ws_tmp.cell(st_y+i,st_x-1).value=col
        for j in range(0,df2.value_counts().size):
            ws_tmp.cell(st_y+i,st_x+j).value=(df2.iloc[j,0]+"\r\n"+str(df2.iloc[j,i]))
            ws_tmp.cell(st_y+i,st_x+j).alignment = Alignment(wrapText=True)
            ws_tmp.column_dimensions[get_column_letter(st_x+j)].width = 15
# 인기 csv 데이터 추출
    st_y=11
    df=df_web_dict["인기검색어"]
    for i in range(0,len(df.columns)):
        col=df.columns[i]
        tmp_col=col.split(">")
        ws_tmp.cell(st_y,st_x+i).value=("\r\n").join(tmp_col[0:len(tmp_col)-1])
        ws_tmp.cell(st_y,st_x+i).alignment = Alignment(wrapText=True)
        ws_tmp.cell(st_y+1,st_x+i).value=tmp_col[len(tmp_col)-1]
        ws_tmp.column_dimensions[get_column_letter(st_x+i)].width = 15
    for x in range(0,len(df.columns)):
        pre=0
        pre_cnt=0
        for i in range(0,len(arr_key_rank)):
            if i>0:
                pre=arr_key_rank[i-1]
                pre_cnt=pre_cnt+arr_key_cnt[i-1]
            ws_tmp.column_dimensions[get_column_letter(9+x)].width = 15
            df2=df.iloc[pre:arr_key_rank[i],x]
            df2=df2.sample(arr_key_cnt[i])
            df2=df2.sort_index()
            for y in range(0,df2.value_counts().size):
                mer_t=str(df2.index[y]+1)+")"+str(df2.iloc[y])
                ws_tmp.cell(st_y+2+y+pre_cnt,st_x+x).value=mer_t

    st_y=st_y+10+3
    df=df_web_dict["인기검색어"]
    for i in range(0,len(df.columns)):
        col=df.columns[i]
        tmp_col=col.split(">")
        ws_tmp.cell(st_y,st_x+i).value=("\r\n").join(tmp_col[0:len(tmp_col)-1])
        ws_tmp.cell(st_y,st_x+i).alignment = Alignment(wrapText=True)
        ws_tmp.cell(st_y+1,st_x+i).value=tmp_col[len(tmp_col)-1]
    for x in range(0,len(df.columns)):
        col=df.columns[x]
        for y in range(0,df.value_counts().size):
            ws_tmp.cell(st_y+2+y,st_x+x).value=str(df.iloc[y][col])

def set_chg_report():
# 광고대비 매출 csv 추출
    df=df_web_dict["광고대비전환"]
    next_y=set_df_to_sht(df,ws_chg,14,2)
    ws=ws_chg;    h=6.5;    w=9.5;    x_l=5;    y1=next_y-4;    y2=next_y;
    x_r=6;    to_cell="K14";
    make_chart(ws,x_l,x_r,y1,y2,to_cell,h,w)
    x_r=7;    to_cell="N14";
    make_chart(ws,x_l,x_r,y1,y2,to_cell,h,w)
    x_r=9;    to_cell="K22";
    make_chart(ws,x_l,x_r,y1,y2,to_cell,h,w)
    x_r=10;    to_cell="N22";
    make_chart(ws,x_l,x_r,y1,y2,to_cell,h,w)

def get_per_text(per,good,norm,bad):
    if per>1:
        return good
    if per==1:
        return norm
    if per<1:
        return bad

def set_desc_report():
    global price_up_cut
    global price_dwn_cut

    month_now=int((datetime.datetime.now()).strftime('%m'))
    year_now=int((datetime.datetime.now()).strftime('%Y'))
    month_next=int((datetime.datetime.now()+dateutil.relativedelta.relativedelta(months=+1)).strftime('%m'))
    month_prev=int((datetime.datetime.now()+dateutil.relativedelta.relativedelta(months=-1)).strftime('%m'))
    year_next=int((datetime.datetime.now()+dateutil.relativedelta.relativedelta(months=+1)).strftime('%Y'))
    year_prev=int((datetime.datetime.now()+dateutil.relativedelta.relativedelta(months=-1)).strftime('%Y'))
    # Desc 총 표기
    df_mas_rpt_sAd=df_web_dict["광고대비전환"].copy().replace(',','', regex=True)
    if not isFirstWeek:
        day_m_l=int((datetime.datetime.now()).strftime('%d'))
        day_m_r=int(calendar.monthrange(year_next,month_next)[1])
        day_chk=(datetime.datetime.now()+dateutil.relativedelta.relativedelta(days=-1)).strftime('(~%m%d)')
        ad_df_tmp=df_mas_rpt_sAd.tail(3)
        month_l=month_now
        month_r=month_next
        rate=20
    else:
        day_m_l=int(calendar.monthrange(year_prev,month_prev)[1])
        day_m_r=int(calendar.monthrange(year_now,month_now)[1])
        day_chk=(date.today().replace(day=1)+dateutil.relativedelta.relativedelta(days=-1)).strftime('(~%m%d)')
        if not str(month_prev)+"월" in df_mas_rpt_sAd.tail(1)["월간"].values[0]:
            ad_df_tmp=df_mas_rpt_sAd.tail(4).head(3)
        else:
            ad_df_tmp=df_mas_rpt_sAd.tail(3)
        month_l=month_prev
        month_r=month_now
        rate=20

    ws_desc["B2"].value=str(month_l)+"월 마케팅 분석"
    ws_desc["I2"].value=str(month_r)+"월 광고집행 의견"
    ws_desc["G4"].value=day_chk

    y=4
    x=2
    for col in ad_df_tmp.columns:
        ws_desc.cell(y,x).value=col
        y=y+1
    for row_idx in range(0,len(ad_df_tmp)):
        x=x+1
        y=4
        ad_df_each=ad_df_tmp.iloc[row_idx]
        for col in ad_df_tmp.columns:
            if col=="월간" or ad_df_each[col]=="-":
                ws_desc.cell(y,x).value=ad_df_each[col]
            elif "%" in str(ad_df_each[col]):
                ws_desc.cell(y,x).value=float(ad_df_each[col].replace("%",""))/100
            else:
                ws_desc.cell(y,x).value=int(ad_df_each[col])
            y=y+1


    path="https://sell.smartstore.naver.com/#/vertical/hotdeal/luckto/list"
    driver_execute(path)
    element = driver.find_element_by_id("__naverpay")
    driver.switch_to.frame(element)
    if len(driver.find_elements_by_xpath("//span[text()='럭투 진행중']"))>0:
        ws_desc["N2"].value="진행중"
        date_up=driver.find_element_by_xpath("//span[text()='럭투 진행중']/parent::td/following-sibling::td[5]").text.split(" ")[0][2:].split("-")
        date_up="~"+date_up[0]+"."+date_up[1]+"/"+date_up[2]
        ws_desc["O2"].value=date_up
    else:
        ws_desc["N2"].value="미진행"
    ws_desc["N2"].hyperlink = path
    ws_desc["N2"].style = "Hyperlink"

    path="https://sell.smartstore.naver.com/#/exposure"
    driver_execute(path)
    blog_ex=driver.find_element_by_xpath('//strong[@ng-class="'+"{'text-primary': vm.NAVERBLOG.hasData}"+'"]').text
    face_ex=driver.find_element_by_xpath('//strong[@ng-class="'+"{'text-primary': vm.FACEBOOK.hasData}"+'"]').text
    insta_ex=driver.find_element_by_xpath('//strong[@ng-class="'+"{'text-primary': vm.INSTAGRAM.hasData}"+'"]').text
    ws_desc["N4"].value=blog_ex
    ws_desc["N5"].value=face_ex
    ws_desc["N6"].value=insta_ex
    if blog_ex=="설정함":
        driver_execute(path)
        blog_link=driver.find_element_by_xpath("//div[@ng-if='vm.NAVERBLOG.url']/dd/span/a").get_attribute("href")
        ws_desc["N4"].hyperlink = blog_link
        ws_desc["N4"].style = "Hyperlink"
        driver_execute(blog_link)
        element = driver.find_element_by_id("mainFrame")
        driver.switch_to.frame(element)
        if len(driver.find_elements_by_xpath("//a[@id='category0']"))>0:
            driver.find_element_by_xpath("//a[@id='category0']").click()
            part_link=driver.find_element_by_xpath("//span[@class='ell2 pcol2']/a").get_attribute("href")
            date_up=driver.find_element_by_xpath("(//span[@class='date pcol2'])[2]").text[2:]
            date_up=date_up.replace(" ","").split(".")
            date_up=date_up[0]+"."+date_up[1]+"/"+date_up[2]
            look_cnt=driver.find_element_by_xpath("//span[@class='num pcol3']").text
        ws_desc["O4"].value=date_up+" "+look_cnt
        ws_desc["O4"].hyperlink = part_link
        ws_desc["O4"].style = "Hyperlink"
    if face_ex=="설정함":
        driver_execute(path)
        face_link=driver.find_element_by_xpath("//div[@ng-if='vm.FACEBOOK.url']/dd/span/a").get_attribute("href")
        ws_desc["N5"].hyperlink = face_link
        ws_desc["N5"].style = "Hyperlink"
        driver_execute(face_link)
        if len(driver.find_elements_by_xpath("//button[@id='loginbutton']"))>0:
            driver.find_element_by_xpath("//input[@id='email']").clear()
            driver.find_element_by_xpath("//input[@id='email']").send_keys("dltjdtn3210@nate.com")
            driver.find_element_by_xpath("//input[@id='pass']").clear()
            driver.find_element_by_xpath("//input[@id='pass']").send_keys("Ss10237209!")
            driver.find_element_by_xpath("//input[@id='pass']").send_keys(Keys.RETURN)
            while len(driver.find_elements_by_xpath("//input[@id='email']"))!=0:
                driver.find_element_by_xpath("//input[@id='email']").clear()
                driver.find_element_by_xpath("//input[@id='email']").send_keys("dltjdtn3210@nate.com")
                driver.find_element_by_xpath("//input[@id='pass']").clear()
                driver.find_element_by_xpath("//input[@id='pass']").send_keys("Ss10237209!")
                driver.find_element_by_xpath("//input[@id='pass']").send_keys(Keys.RETURN)
                time.sleep(1)
            driver_execute(face_link)
            first_ele=driver.find_element_by_xpath("//div[@aria-posinset='1']")
            posi_m=first_ele.find_element_by_xpath("//span[contains(@id,'jsc_c')]")
            action = ActionChains(driver)
            action.move_to_element(posi_m).perform()
            date_up=first_ele.find_element_by_xpath("//div[@class='__fb-light-mode']").text[2:].split(" ")
            while len(date_up)==1:
                action.move_to_element(posi_m).perform()
                date_up=driver.find_element_by_xpath("//div[@class='__fb-light-mode']").text[2:].split(" ")
            date_up="".join(date_up[:3])
            date_up=date_up.replace("년",".").replace("월","/").replace("일","")
            look_cnt=driver.find_element_by_xpath("//span[@class='gpro0wi8 pcp91wgn']").text
        else:
            date_up=driver.find_element_by_xpath("//abbr").get_attribute("title")[2:].split(" ")
            date_up="".join(date_up[:3])
            date_up=date_up.replace("년",".").replace("월","/").replace("일","")
            look_cnt=driver.find_element_by_xpath("//span[@class='_81hb']").text
        ws_desc["O5"].value=date_up+" ("+look_cnt+"회)"
    if insta_ex=="설정함":
        driver_execute(path)
        insta_link=driver.find_element_by_xpath("//div[@ng-if='vm.INSTAGRAM.url']/dd/span/a").get_attribute("href")
        ws_desc["N6"].hyperlink = insta_link
        ws_desc["N6"].style = "Hyperlink"
        driver_execute(insta_link)
        if len(driver.find_elements_by_xpath("//span[text()='Facebook으로 로그인']"))>0:
            driver.find_element_by_xpath("//span[text()='Facebook으로 로그인']").click()
            driver.find_element_by_xpath("//input[@id='email']").clear()
            driver.find_element_by_xpath("//input[@id='email']").send_keys("dltjdtn3210@nate.com")
            driver.find_element_by_xpath("//input[@id='pass']").clear()
            driver.find_element_by_xpath("//input[@id='pass']").send_keys("Ss10237209!")
            driver.find_element_by_xpath("//input[@id='pass']").send_keys(Keys.RETURN)
            driver.find_element_by_xpath("//button[text()='나중에 하기']").click()
            driver_execute(insta_link)
        elif len(driver.find_elements_by_xpath("//div[contains(text(),'계속')]"))>0:
            driver.find_element_by_xpath("//div[contains(text(),'계속')]").click()
            driver.find_element_by_xpath("//nav")
            driver_execute(insta_link)
        part_link=driver.find_element_by_xpath("//div[@style='flex-direction: column; padding-bottom: 0px; padding-top: 0px;']/div/div/a").get_attribute("href")
        driver_execute(part_link)
        date_up=driver.find_element_by_xpath("//time").get_attribute("title")[2:].replace("년 ",".").replace("월 ","/").replace("일","")
        look_cnt=driver.find_element_by_xpath("//span[contains(text(),'조회')]/span").text
        ws_desc["O6"].value=date_up+" ("+look_cnt+"회)"
        ws_desc["O6"].hyperlink = part_link
        ws_desc["O6"].style = "Hyperlink"

    cur_ad=int(ws_desc["E7"].value)
    ad_per=round(int(ws_desc["E7"].value)/((int(ws_desc["C7"].value)+int(ws_desc["D7"].value))/2),2)
    ad_text=get_per_text(ad_per,str(ad_per)+"배 증가","현상 유지",str(ad_per)+"배 감소")
    buy1_per=round(int(ws_desc["E8"].value)/((int(ws_desc["C8"].value)+int(ws_desc["D8"].value))/2),2)
    buy1_text=get_per_text(buy1_per,str(buy1_per)+"배 증가","현상 유지",str(buy1_per)+"배 감소")
    buy2_per=round(int(ws_desc["E9"].value)/((int(ws_desc["C9"].value)+int(ws_desc["D9"].value))/2),2)
    buy2_text=get_per_text(buy2_per,str(buy2_per)+"배 증가","현상 유지",str(buy2_per)+"배 감소")
    to_ad_per=round(float(ws_desc["E12"].value)/((float(ws_desc["C12"].value)+float(ws_desc["D12"].value))/2),2)
    to_ad_text=get_per_text(to_ad_per,str(to_ad_per)+"배 좋아졌습니다.","증감이 없습니다.",str(to_ad_per)+"배 나빠졌습니다.")
    ef_per=round(float(ws_desc["E10"].value)/((float(ws_desc["C10"].value)+float(ws_desc["D10"].value))/2),2)
    ef_text=get_per_text(ef_per,str(ef_per)+"배 좋아졌습니다.","양호합니다.",str(ef_per)+"배 나빠졌습니다.")
    ef_text2=""
    total_per=round(int(ws_desc["E11"].value)/((int(ws_desc["C11"].value)+int(ws_desc["D11"].value))/2),2)
    total_text=get_per_text(total_per,str(total_per)+"배 증가","현상 유지",str(total_per)+"배 감소")
    total_desc=get_per_text(total_per,"성장세를 보이고 있습니다.","추이가 안정적입니다.","SNS나 럭키투데이 등 마케팅이 필요해보입니다.")

    ws_desc["I4"].value=str(month_r)+"월필요비용"
    ws_desc["I5"].value=str(month_l)+"월잔여금액"

    ad_need=int(ws_desc["E7"].value)*ef_per*day_m_r/day_m_l
    ws_desc["K4"].value=int(round(ad_need))
    ws_desc["K5"].value=int(df_admin_dict["광고잔액"])
    price_up_cut=int(df_admin_dict['입찰가하강순위(이하)'])
    price_dwn_cut=int(df_admin_dict['입찰가상승순위(이상)'])
    ws_desc["K6"].value=str(price_up_cut)+"~"+str(price_dwn_cut)+"위"

    if price_up_cut>1 and ef_per>1:
        price_up_cut=price_up_cut-1
        price_dwn_cut=price_dwn_cut-1
        ef_text2="따라서, 광고노출순위를 "+str(price_up_cut)+"~"+str(price_dwn_cut)+"위로 올립니다."
    elif ef_per<1:
        price_up_cut=price_up_cut+1
        price_dwn_cut=price_dwn_cut+1
        ef_text2="따라서, 광고노출순위를 "+str(price_up_cut)+"~"+str(price_dwn_cut)+"위로 내립니다."
    elif price_up_cut==1:
        ef_text2="광고노출순위는 더 이상 올릴 수 없으므로 유지합니다."
    else:
        ef_text2="따라서, 광고노출순위는 "+str(price_up_cut)+"~"+str(price_dwn_cut)+"위로 유지합니다."
    need_ad_price="약 "+str(round(float(ad_need)/10000,1)).replace(".0","")+"만원 정도 필요할 것으로 보입니다."

    ws_desc["I7"].value="ROAS율 추이가 "+ef_text
    ws_desc["I8"].value=ef_text2
    ws_desc["I9"].value="(광고비 "+ad_text+" 대비, 전환수 "+buy1_text+" / 전환액 "+buy2_text+")"
    ws_desc["I10"].value=str(month_r)+"월 광고비는 "+need_ad_price
    ws_desc["I11"].value="매출총액은 "+total_text+"로 "+total_desc
    ws_desc["I12"].value=str(month_l)+"월 매출의 "+str(round(ws_desc["E12"].value*100))+"%는 광고 유입이며, 3개월 평균 대비 "+ to_ad_text

    df_mas_sum=df_web_dict["소재부분합리스트"].copy()[['대표이미지 URL','기본상품명','상품가격','네이버쇼핑 카테고리','노출수','클릭수','전환수','쇼핑몰 상품 ID','전환매출액(원)']]
    df_mas_worst=df_mas_sum[df_mas_sum["노출수"] > df_mas_sum["노출수"].quantile((100-rate)/100)].copy()
    df_mas_worst=df_mas_worst.sort_values(by="클릭수", ascending=True).head(2)
    df_mas_best=df_mas_sum[(df_mas_sum["전환매출액(원)"] > df_mas_sum["전환매출액(원)"].quantile((100-rate)/100)) & df_mas_sum["전환매출액(원)"]!=0].copy()
    df_mas_best=df_mas_best.sort_values(by="전환매출액(원)", ascending=False).head(1)
    df_mas_best_cnt=df_mas_sum[(df_mas_sum["전환수"] > df_mas_sum["전환수"].quantile((100-rate)/100)) & df_mas_sum["전환매출액(원)"]!=0].copy()
    df_mas_best_cnt=df_mas_best_cnt.sort_values(by="전환수", ascending=False).head(2)
    for i in range(0,len(df_mas_best_cnt)):
        df_tmp=df_mas_best_cnt.iloc[i]["쇼핑몰 상품 ID"]
        if df_tmp!=df_mas_best.iloc[0]["쇼핑몰 상품 ID"]:
            df_mas_best=df_mas_best.append(df_mas_best_cnt.iloc[i])
            break

    y=17
    x=4
    for idx in range(0,len(df_mas_best)):
        each_mas=df_mas_best.iloc[idx]
        insert_img(each_mas['대표이미지 URL'],y,x-2,ws_desc)
        insert_img(each_mas['대표이미지 URL'],y-12,x-2,ws_ad)
        ws_desc.cell(y,x).value=each_mas['기본상품명']
        ws_desc.cell(y,x).hyperlink = "https://smartstore.naver.com/signcody/products/"+str(each_mas['쇼핑몰 상품 ID'])
        ws_desc.cell(y,x).style = "Hyperlink"
        ws_desc.cell(y+1,x).value=each_mas['상품가격']
        ws_desc.cell(y+2,x).value=each_mas['네이버쇼핑 카테고리']
        ws_desc.cell(y,x+2).value=each_mas['쇼핑몰 상품 ID']
        ws_desc.cell(y+1,x+3).value=each_mas['노출수']
        ws_desc.cell(y+1,x+4).value=each_mas['클릭수']
        ws_desc.cell(y+1,x+5).value=each_mas['전환수']
        ws_desc.cell(y+1,x+6).value=each_mas['전환매출액(원)']
        feed_title="-"
        feed_detail="-"
        if not isFirstWeek:
            warn_dict={
                "연관성 있는 카테고리":    "카테고리수정 요청/상품과 연관성 있는 카테고리로 수정해 주세요",
                "분할된 비율":             "이미지수정 요청/분할된 이미지는 서로 달라야 하며 분할비율이 동일해야 합니다.",
                "이미지 내 텍스트":        "이미지수정 요청/이미지 내 텍스트가 기재된 경우 광고등록이 불가합니다.",
                "테두리, 공백이 확인":     "이미지수정 요청/공백, 테두리가 눈에 띄거나, 저화질 이미지는 광고진행이 불가합니다.",
                "무의미하게 반복나열":     "상품명수정 요청/상품과 관련없는 수식어를 나열하거나 유사문구를 반복 기재할 수 없습니다."
            }
            df_mas_warn=df_web_dict["☆점검소재리스트"].copy()[['쇼핑몰 상품 ID','기본상품명','노출상품명','제한 사유']]
            df_warn_each=df_mas_warn[(df_mas_warn["쇼핑몰 상품 ID"].astype(str)==each_mas['쇼핑몰 상품 ID'].astype(str))]
            if len(df_warn_each)>0:
                df_warn_each=df_warn_each.iloc[0]
                for key in warn_dict.keys():
                    if df_warn_each["제한 사유"].find(key)>-1:
                        feed_title=warn_dict[key].split("/")[0]
                        feed_detail=warn_dict[key].split("/")[1]
                        break
        ws_desc.cell(y+1,x+2).value=feed_title
        ws_desc.cell(y+2,x+2).value=feed_detail
        y=y+3
    y=24
    x=4
    for idx in range(0,len(df_mas_worst)):
        each_mas=df_mas_worst.iloc[idx]
        insert_img(each_mas['대표이미지 URL'],y,x-2,ws_desc)
        insert_img(each_mas['대표이미지 URL'],y-12,x-2,ws_ad)
        ws_desc.cell(y,x).value=each_mas['기본상품명']
        ws_desc.cell(y,x).hyperlink = "https://smartstore.naver.com/signcody/products/"+str(each_mas['쇼핑몰 상품 ID'])
        ws_desc.cell(y,x).style = "Hyperlink"
        ws_desc.cell(y+1,x).value=each_mas['상품가격']
        ws_desc.cell(y+2,x).value=each_mas['네이버쇼핑 카테고리']
        ws_desc.cell(y,x+2).value=each_mas['쇼핑몰 상품 ID']
        ws_desc.cell(y+1,x+3).value=each_mas['노출수']
        ws_desc.cell(y+1,x+4).value=each_mas['클릭수']
        ws_desc.cell(y+1,x+5).value=each_mas['전환수']
        ws_desc.cell(y+1,x+6).value=each_mas['전환매출액(원)']
        feed_title="-"
        feed_detail="-"
        if not isFirstWeek:
            warn_dict={
                "연관성 있는 카테고리":    "카테고리수정 요청/상품과 연관성 있는 카테고리로 수정해 주세요",
                "분할된 비율":             "이미지수정 요청/분할된 이미지는 서로 달라야 하며 분할비율이 동일해야 합니다.",
                "이미지 내 텍스트":        "이미지수정 요청/이미지 내 텍스트가 기재된 경우 광고등록이 불가합니다.",
                "테두리, 공백이 확인":     "이미지수정 요청/공백, 테두리가 눈에 띄거나, 저화질 이미지는 광고진행이 불가합니다.",
                "무의미하게 반복나열":     "상품명수정 요청/상품과 관련없는 수식어를 나열하거나 유사문구를 반복 기재할 수 없습니다."
            }
            df_mas_warn=df_web_dict["☆점검소재리스트"].copy()[['쇼핑몰 상품 ID','기본상품명','노출상품명','제한 사유']]
            df_warn_each=df_mas_warn[(df_mas_warn["쇼핑몰 상품 ID"].astype(str)==each_mas['쇼핑몰 상품 ID'].astype(str))]
            if len(df_warn_each)>0:
                df_warn_each=df_warn_each.iloc[0]
                for key in warn_dict.keys():
                    if df_warn_each["제한 사유"].find(key)>-1:
                        feed_title=warn_dict[key].split("/")[0]
                        feed_detail=warn_dict[key].split("/")[1]
                        break
            else:
                feed_title="이미지 수정 요청"
                feed_detail="상품성 있는 이미지로 수정이 필요해보입니다."
        ws_desc.cell(y+1,x+2).value=feed_title
        ws_desc.cell(y+2,x+2).value=feed_detail
        y=y+3

    df=df_web_dict["검색어-보고서"][["검색어","노출수","클릭수","전환수"]]
    df=df[1:]
    columnNames = df.columns
    st_y=17
    for i in range(1,4):
        col=columnNames[i]
        df[col] = df[col].astype(str).str.replace(',', '')
        df=df.astype({col: int})
        df=df.sort_values(by=col, ascending=False)
        df2=df[df[col] > df[col].quantile((100-10)/100)]
        df2=df2.head(3)["검색어"]
        ws_desc.cell(st_y+(i-1)*3,12).value=", ".join(df2.tolist())+" 키워드는"

def set_df_admin(camp_id,ch_col,val):
    file_admin=find_file("※광고주관리",THIS_FOLDER)
    if file_admin!="False":
        admin_file=os.path.join(THIS_FOLDER,file_admin)
        df_admin=pd.read_csv(admin_file, encoding='utf-8-sig')
        df_admin.loc[df_admin["CAMP_ID"]==camp_id,ch_col]=val
        df_admin.to_csv(THIS_FOLDER+"/※광고주관리.csv", encoding='utf-8-sig', index=False)
    return df_admin

def set_ad_report():
    df_all=df_web_dict["소재부분합리스트"].copy()
    df_all_sum=df_all.sum()[["노출수","클릭수","전환수","총비용(VAT포함,원)","전환매출액(원)"]]
    df_all_sum.loc["클릭률"]=(df_all_sum["클릭수"]*10000/df_all_sum["노출수"]//1)/100
    df_all_sum.loc["전환율"]=(df_all_sum["전환수"]*10000/df_all_sum["클릭수"]//1)/100
    y=20
    x=2
    group_start_x=0
    group_last_x=1
    for total_val in df_all.columns:
        group_last_x=group_last_x+1
        if total_val in df_all_sum.index:
            ws_ad.cell(y-1,x).value=df_all_sum[total_val]
        ws_ad.cell(y,x).value=total_val
        if "소재1" == total_val:
            group_start_x=group_last_x
            v_idx=0
            tmp_df_val=""
            for each_col in report_col_val:
                tmp_df_val = tmp_df_val + each_col
                tmp_df_val = tmp_df_val + report_col_val_del[v_idx]
                v_idx=v_idx+1
                ws_ad.cell(y-1,x).value=tmp_df_val+"\r\n(노출▶클릭순▶)"
                ws_ad.cell(y-1,x).alignment = Alignment(wrapText=True)
                # ws_ad.row_dimensions[y-1].height = 115
            yy=16
            xx=x
            for gg in gg_val:
                ws_ad.cell(yy,xx).value=gg
                ws_ad.cell(yy,xx).fill=PatternFill("solid", fgColor=gg_col[gg_val.index(gg)])
                if yy==17:
                    yy=16
                    xx=xx+1
                else:
                    yy=yy+1
        x=x+1
    y=y+1
    for col_idx in range(group_start_x,group_last_x+1):
        ws_ad.column_dimensions.group(colnum_string(col_idx), hidden=True)

    for df_y in range(0,len(df_all)):
        x=2
        for df_x in range(0,len(df_all.iloc[df_y])):
            ws_ad.cell(y,x).value=df_all.iloc[df_y,df_x]
            if "기본상품명" in df_all.columns[df_x]:
                ws_ad.cell(y,x).hyperlink = "https://smartstore.naver.com/signcody/products/"+str(df_all["쇼핑몰 상품 ID"].iloc[df_y])
                ws_ad.cell(y,x).style = "Hyperlink"
            elif "대표이미지 URL" in df_all.columns[df_x]:
                ws_ad.column_dimensions[colnum_string(x)].width = 7.5
                url = df_all["대표이미지 URL"].iloc[df_y]
                insert_img(url,y,x,ws_ad)
            elif "소재" in df_all.columns[df_x]:
                ws_ad.column_dimensions[colnum_string(x)].width = 30
                ws_ad.row_dimensions[y].height = 48
                ws_ad.cell(y,x).alignment = Alignment(wrapText=True)
                delim=str(ws_ad.cell(y,x).value).split("\r\n")[0]
                for gg in gg_val:
                    gg_idx=gg_val.index(gg)
                    if gg in delim:
                        ws_ad.cell(y,x).fill=PatternFill("solid", fgColor=gg_col[gg_idx])
            x=x+1
        y=y+1

def insert_img(url,y,x,sht):
    http = urllib3.PoolManager()
    r = http.request('GET', url)
    image_file = io.BytesIO(r.data)
    img = Image(image_file)
    size_b=sht.column_dimensions[colnum_string(x)].width*7.98
    img.width= size_b
    img.height = size_b
    cell_posi=get_column_letter(x)+str(y)
    img.anchor = cell_posi
    sht.add_image(img)

def driver_logout():
    global login_status
    if login_status:
        if store_id_exist:
            time.sleep(1)
            pre_logout("https://sell.smartstore.naver.com/#/logout")
        time.sleep(1)
        pre_ad_logout("https://searchad.naver.com/logout")
        login_status=False

def driver_execute(path):
    global driver_exist
    global login_status
    if not driver_exist:
        driver_init()
        driver_exist=True
    if not login_status:
        login_status=True
        time.sleep(1)
        if store_id_exist:
            pre_login("https://sell.smartstore.naver.com/#/bizadvisor/marketing")
        time.sleep(1)
        pre_ad_login("https://manage.searchad.naver.com/customers/"+ str(df_admin_dict["CUST_ID"])+"/reports/"+ str(df_admin_dict["검색어보고서"]))
        get_ad_money("https://manage.searchad.naver.com/customers/"+str(df_admin_dict["CUST_ID"])+"/billing/bizmoney")

    if driver.current_url!=path:
        time.sleep(1)
        driver.get(path)

def driver_init():
    global driver

    g_name=os.path.join(THIS_FOLDER,"chromedriver")
    options = webdriver.ChromeOptions()
    # 창 숨기는 옵션 추가
    # options.add_argument("headless")
    prefs = {
      "download.default_directory": FILE_FOLDER,
      "download.prompt_for_download": False,
      "download.directory_upgrade": True,
      "safebrowsing.enabled": True,
      "profile.default_content_setting_values.notifications": 1
    }
    options.add_argument("--start-maximized")
    options.add_experimental_option('prefs', prefs)

    driver = webdriver.Chrome(g_name, options=options)
    driver.implicitly_wait(10)

# ---------------- 서브 모듈
def timer_start():
    global start_t
    start_t = time.time()

def timer_chk():
    print("time :", time.time() - start_t)

def find_file(fname,FOLDER):
    file_flist=os.listdir(FOLDER)
    for file_f in file_flist:
        if file_f.find(fname)>-1:
            return file_f
    return "False"

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)

def make_dummydf(df,fname):
    df.to_csv(FILE_FOLDER+'/'+fname+'.csv' , encoding='utf-8-sig',index = False)

def test():
    while True:
        clk=input("c or e or any~~")
        try:
            print(clk)
            if clk=="c":
                key=input("key")
                driver.find_element_by_xpath(key).click()
            elif clk=="e":
                break
            elif clk=="m":
                key=input("key")
                posi_m=driver.find_element_by_xpath(key)
                action = ActionChains(driver)
                action.move_to_element(posi_m).perform()
            else:
                key=input("key")
                print(driver.find_element_by_xpath(key).get_attribute("innerHTML"))
        except Exception as e:
            print("Error:" + str(e))

def generate(timestamp, method, uri, secret_key):
    message = "{}.{}.{}".format(timestamp, method, uri)
#     hash = hmac.new(bytes(secret_key, "utf-8"), bytes(message, "utf-8"), hashlib.sha256)
    hash = hmac.new(secret_key.encode("utf-8"), message.encode("utf-8"), hashlib.sha256)
    hash.hexdigest()
    return base64.b64encode(hash.digest())

def get_header(method, uri, api_key, secret_key, customer_id):
    timestamp = str(int(time.time() * 1000))
    signature = generate(timestamp, method, uri, SECRET_KEY)
    return {'Content-Type': 'application/json; charset=UTF-8', 'X-Timestamp': timestamp, 'X-API-KEY': API_KEY, 'X-Customer': str(CUSTOMER_ID), 'X-Signature': signature}

def call_RelKwd(_kwds_string):
    global BASE_URL,CUSTOMER_ID,API_KEY,SECRET_KEY
    BASE_URL = 'https://api.naver.com'
    CUSTOMER_ID = '392590'
    API_KEY = '01000000008fa2a584355277148cf5b1792f0a1650becf66b049812186ce69dbfb7cbf4ec3'
    SECRET_KEY = 'AQAAAACPoqWENVJ3FIz1sXkvChZQySFY24NP0GvvyT3R14cXaQ=='

    uri = '/keywordstool'
    method = 'GET'
    prm = {'hintKeywords' : _kwds_string , 'showDetail':1}
    # ManageCustomerLink Usage Sample
    returnData = None
    df = pd.DataFrame()
    # print(_kwds_string)
    repeat_time=0.5
    wait_time=3
    while returnData is None:
        try:
            r = requests.get(BASE_URL + uri, params=prm, headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
            returnData = r.json()
            df = pd.DataFrame(returnData['keywordList'])
        except Exception as e:
            if 'code' in returnData:
                time.sleep(wait_time)
                wait_time=wait_time+1
                print(_kwds_string)
                print(wait_time)
            time.sleep(repeat_time)
            r = requests.get(BASE_URL + uri, params=prm, headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
            returnData = r.json()
            df = pd.DataFrame(returnData['keywordList'])
            pass
        if not 'keywordList' in returnData:
            print(_kwds_string)
            print(returnData)
            return False

    # df.to_csv(FILE_FOLDER+"/sdfasdf.csv", encoding='utf-8-sig')
    df['총검색']=df['monthlyMobileQcCnt'].astype(str).replace("< 10",0).astype(float)+df['monthlyPcQcCnt'].astype(str).replace("< 10",0).astype(float)
    df['총클릭']=df['monthlyAveMobileClkCnt'].astype(float)+df['monthlyAvePcClkCnt'].astype(float)
    df.rename({'compIdx':'경쟁도',       'monthlyAveMobileClkCnt':'평균클릭(폰)',       'monthlyAveMobileCtr':'평균클릭률(폰)',
       'monthlyAvePcClkCnt':'평균클릭(PC)',       'monthlyAvePcCtr':'평클릭률(PC)',       'monthlyMobileQcCnt':'검색(폰)',
       'monthlyPcQcCnt': '검색(PC)',       'plAvgDepth':'노출광고수',       'relKeyword':'연관키워드'},axis=1,inplace=True)
    rate=0.95
    if len(df)>1:
        tmp_df=df[(df['총검색'] > df['총검색'].quantile(rate)) & (df['총클릭'] > df['총클릭'].quantile(rate))]
        while len(tmp_df)==0:
            tmp_df=df[(df['총클릭'] > df['총클릭'].quantile(rate))]
            rate=rate-0.05
        df=tmp_df
        df=df.sort_values(by="총클릭", ascending=False)
    return df

def get_cate_key(path):
    driver_execute(path)
    cate_df=df_web_dict["광고소재리스트"]["네이버쇼핑 카테고리"].copy().dropna().drop_duplicates().values
    cate_my=df_web_dict["광고소재리스트"]["네이버쇼핑 카테고리"].copy().dropna().drop_duplicates().tolist()
    cate_all_file=find_file("인기검색어.csv",THIS_FOLDER)
    if find_file(cate_all_file,THIS_FOLDER)!="False":
        cate_all_file_name=os.path.join(THIS_FOLDER,cate_all_file)
        df_cate_all=pd.read_csv(cate_all_file_name, encoding='utf-8-sig')
        df_cate_all_col=df_cate_all.columns.tolist()
        cate_my=np.setdiff1d(cate_my,df_cate_all_col)
    else:
        df_cate_all=pd.DataFrame()
    for cate in cate_my:
        cate_arr=cate.split(">")
        driver.find_element_by_xpath('//*[@id="18_device_0"]').click()
        driver.find_element_by_xpath('//*[@id="19_gender_0"]').click()
        driver.find_element_by_xpath('//*[@id="20_age_0"]').click()
        for i in range(0, len(cate_arr)):
            driver.find_element_by_xpath("(//span[contains(@class,'select_btn')])["+str(i+1)+"]").click()
            driver.find_element_by_xpath("(//a[text()='"+cate_arr[i].strip()+"'])").click()
        driver.find_element_by_xpath('//*[@id="content"]/div[2]/div/div[1]/div/a').click()
        time.sleep(1)
        arr_pro=[]
        for p in range(0, 25):
            for i in range(1, 21):
                keyword_path = '//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[1]/ul/li[{}]/a'.format(i)
                key_num=driver.find_element_by_xpath(keyword_path).text.split("\n")[0]
                key_word=driver.find_element_by_xpath(keyword_path).text.split("\n")[1].replace(" ","")
                while(int(key_num)!=p*20+i):
                    time.sleep(0.1)
                    keyword_path = '//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[1]/ul/li[{}]/a'.format(i)
                    key_num=driver.find_element_by_xpath(keyword_path).text.split("\n")[0]
                    key_word=driver.find_element_by_xpath(keyword_path).text.split("\n")[1]
                if len(arr_pro)<int(key_num)+1:
                    arr_pro.append(key_word)
                else:
                    arr_pro[int(key_num)].append(key_word)
            driver.find_element_by_xpath('//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/a[2]').click()
        df_cate_all[cate]=arr_pro
    if nowWeekDay=="일":
        cate_all_file=weekHold+"인기검색어.csv"
    else:
        cate_all_file=weekDel+"인기검색어.csv"
    cate_all_file_name=os.path.join(THIS_FOLDER,cate_all_file)
    df_cate_all.to_csv(cate_all_file_name, encoding='utf-8-sig', index = False)
    return df_cate_all[cate_df]
    # return keyword_list

def get_rel_key(each_show,df_pid):
    each_show=each_show.replace(",","").replace(".","").replace("/","").replace("(","").replace(")","").replace("-","")
    each_show=' '.join(each_show.split())
    rel_all_file=find_file("연관검색어.csv",THIS_FOLDER)
    if find_file(rel_all_file,THIS_FOLDER)!="False":
        rel_all_file_name=os.path.join(THIS_FOLDER,rel_all_file)
        df_rel_key=pd.read_csv(rel_all_file_name, encoding='utf-8-sig')
    else:
        df_rel_key=pd.DataFrame(columns=(call_RelKwd("test").columns.insert(0,"노출키워드")))
    df_rel_key_arr=df_rel_key["노출키워드"].drop_duplicates().tolist()
    tmp_each_show=str(each_show).split(" ")
    for each_show_2 in tmp_each_show:
        if not each_show_2 in df_rel_key_arr and not each_show_2 in [",", ".", "/", "(", ")", "", "-"]:
            df_rel_key_arr.append(each_show_2)
            df_rel_key_each=call_RelKwd(str(each_show_2))
            if isinstance(df_rel_key_each, pd.DataFrame):
                df_rel_key_each['노출키워드']=each_show_2
                df_rel_key=df_rel_key.append(df_rel_key_each, ignore_index = True)
                if nowWeekDay=="일":
                    rel_all_file=weekHold+"연관검색어.csv"
                else:
                    rel_all_file=weekDel+"연관검색어.csv"
                rel_all_file_name=os.path.join(THIS_FOLDER,rel_all_file)
                df_rel_key.to_csv(rel_all_file_name, encoding='utf-8-sig', index = False)
    return make_new_name(each_show,df_rel_key,df_pid)

def make_new_name(old_name,df_rel_key,df_pid):
# 연관검색어 교체
    key_arr=old_name.split(" ")
    new_arr=[]
    for key in key_arr:
        tmp_arr=df_rel_key.loc[df_rel_key["노출키워드"]==key,"연관키워드"].values.tolist()
        tmp_idx=random.randint(0, len(tmp_arr)-1)
        tmp_key=tmp_arr[tmp_idx]
        new_arr.append(tmp_key)
        tmp_name=trim_name(" ".join(new_arr))
        while len(tmp_name)>25:
            new_arr.pop(len(new_arr)-1)
            tmp_name=trim_name(" ".join(new_arr))
    new_name=" ".join(new_arr)
    print("1",old_name)
    print("2",new_name)

# 인기검색어 추가
    pop_all_file=find_file("상품리스트.csv",FILE_FOLDER)
    if find_file(pop_all_file,FILE_FOLDER)!="False":
        pop_all_file_name=os.path.join(FILE_FOLDER,pop_all_file)
        df_web_key=pd.read_csv(pop_all_file_name, encoding='utf-8-sig')
    df_web_key=df_web_key[["상품번호(스마트스토어)","카테고리명"]].copy().drop_duplicates()
    df_web_key=df_web_key.loc[df_web_key["상품번호(스마트스토어)"].astype(str)==str(df_pid),"카테고리명"].astype(str).str.replace(">"," > ").values[0]
    pop_cate_file=find_file("인기검색어.csv",THIS_FOLDER)
    if find_file(pop_cate_file,THIS_FOLDER)!="False":
        pop_cate_file_name=os.path.join(THIS_FOLDER,pop_cate_file)
        cate_web_key=pd.read_csv(pop_cate_file_name, encoding='utf-8-sig')
    key_arr=cate_web_key[df_web_key].copy().head(25).values.tolist()
    while len(new_name)<25:
        key_idx=random.randint(0, len(key_arr)-1)
        tmp_key=key_arr[key_idx][0]
        new_arr.append(tmp_key)
        tmp_name=trim_name(" ".join(new_arr))
        if len(tmp_name)>25:
            break
        new_name=tmp_name
    return new_name

def trim_name(name):
    tmp_str=name.split(" ")
    for cc in range(0,len(tmp_str)):
        tmp2=tmp_str[cc]
        for cc2 in range(len(tmp2),1):
            tmp3=tmp2[0:cc2]
            for cc3 in range(cc+1,len(tmp_str)):
                if tmp_str[cc3].find(tmp3)>-1:
                    tmp_str[cc3]=tmp_str[cc3].replace(tmp3,"")
    for cc in range(len(tmp_str)-1,-1):
        tmp2=tmp_str[cc]
        for cc2 in range(1,len(tmp2)-1):
            tmp3=tmp2[cc2:]
            for cc3 in range(cc-1,-1):
                if tmp_str[cc3].find(tmp3)>-1:
                    tmp_str[cc3]=tmp_str[cc3].replace(tmp3,"")
    for cc in range(0,len(tmp_str)-1):
        tmp2=tmp_str[cc]
        for cc2 in range(1,len(tmp2)-1):
            tmp3=tmp2[cc2:len(tmp2)]
            for cc3 in range(cc+1,len(tmp_str)):
                if tmp_str[cc3].find(tmp3)>-1:
                    tmp_str[cc3]=tmp_str[cc3].replace(tmp3,"")

    print("3",name)
    print("4"," ".join(" ".join(tmp_str).split()).strip())
    return " ".join(" ".join(tmp_str).split()).strip()

def pre_logout(path):
    driver_execute(path)

def pre_ad_logout(path):
    driver_execute(path)

def pre_login(path):
    driver_execute(path)
    driver.find_element_by_xpath("//input[@id='loginId']").clear()
    driver.find_element_by_xpath("//input[@id='loginId']").send_keys("signcody4657")
    driver.find_element_by_xpath("//input[@id='loginPassword']").clear()
    driver.find_element_by_xpath("//input[@id='loginPassword']").send_keys("dsp3134233")
    driver.find_element_by_xpath("//input[@id='loginPassword']").send_keys(Keys.RETURN)

def pre_ad_login(path):
    driver_execute(path)
    driver.find_element_by_xpath("//a[@class='btn btn-lg btn-block btn-primary']").click()
    driver.find_element_by_xpath("//input[@id='uid']").clear()
    driver.find_element_by_xpath("//input[@id='uid']").send_keys("signcody")
    driver.find_element_by_xpath("//input[@id='upw']").clear()
    driver.find_element_by_xpath("//input[@id='upw']").send_keys("dsp3134233")
    driver.find_element_by_xpath("//input[@id='upw']").send_keys(Keys.RETURN)

def get_ad_money(path):
    driver_execute(path)
    ad_biz_money=driver.find_element_by_xpath("//span[@class='amount']").text.replace(",","").replace("원","")
    df_admin_dict["광고잔액"]=ad_biz_money

def click_element_by_xpath(xpath):
    action = ActionChains(driver)
    action.move_to_element(driver.find_element_by_xpath(xpath)).perform()
    driver.find_element_by_xpath(xpath).click()

def get_text_by_xpath(xpath):
    tmp_df=""
    while tmp_df=="":
        try:
            tmp_df=driver.find_element_by_xpath(xpath).text
        except Exception as e:
            print(e)
    return tmp_df

def get_pr_report(path):
    driver_execute(path)
    element = driver.find_element_by_id("__naverpay")
    driver.switch_to.frame(element)
    time.sleep(1)
    arr_col=["월간"]
    arr_val=[]
    tmp_arr=driver.find_elements_by_xpath("//thead/tr[2]/th[@scope='col']")
    for eachc in tmp_arr:
        arr_col.append(eachc.text.replace("\n",""))
    time.sleep(1)
    driver.find_element_by_xpath("//a[@class='btn select_data']").click()
    print(nowDate[6:])
    if nowDate[6:]=="01":
        driver.find_element_by_xpath("//span[text()='지난 달']").click()
    else:
        driver.find_element_by_xpath("//span[text()='이번 달']").click()
    driver.find_element_by_xpath("//span[text()='적용']").click()
    while True:
        tmp_date=driver.find_element_by_xpath("//a[@class='btn select_data']/span/span/em").get_attribute("innerHTML")
        tmp_y=tmp_date.split(".")[0].strip()
        tmp_m=tmp_date.split(".")[1].strip()

        tmp_div=[tmp_y[2:]+"."+tmp_m+"월"]
        val_div=driver.find_elements_by_xpath("//tr[@class='total_row']/td")
        for i in range(2,len(val_div)):
            tmp_div.append(val_div[i].text.replace(",",""))
        arr_val.append(tmp_div)

        driver.find_element_by_xpath("//a[@class='btn select_data']").click()
        date_div=driver.find_elements_by_xpath("//div[@class='DayPicker-Month'][2]/div/div/div[@aria-disabled='false']")
        if len(date_div)==0:
            break
        date_div[0].click()
        prev_date=get_text_by_xpath("//div[@class='DayPicker-Month'][2]/div")
        tmp_y=int(prev_date.split(".")[0])
        tmp_m=int(prev_date.split(".")[1])
        prev_arr=calendar.monthrange(tmp_y,tmp_m)
        last_part=datetime.date(tmp_y,tmp_m,prev_arr[1]).strftime("%a %b %d %Y")
        driver.find_element_by_xpath("//div[@aria-label='"+last_part+"']").click()
        driver.find_element_by_xpath("//span[text()='적용']").click()
    df = pd.DataFrame(arr_val)
    df.columns = arr_col
    df=df.drop(columns=['모바일비율(결제금액)', '결제당결제금액', '환불비율(결제금액)'])
    df=df.sort_values(by=['월간'], ascending=[True])
    return df

def get_adp_report(path):
    driver_execute(path)
    time.sleep(1)
    arr_pro=[]
    from_col=[      '노출수',   '클릭수', '클릭률(%)',  '총비용(VAT포함,원)', '전환수', '전환율(%)', '전환매출액',  '광고수익률(%)']
    arr_col=['월간', '노출수',  '클릭수' ,'클릭률',    '광고비',             '전환수', '전환율',    '전환액',      'ROAS율']
    idx=0
    while True:
        if idx==0:
            driver.find_element_by_xpath("//div[@style='background: rgb(255, 255, 255); color: rgb(51, 51, 51); line-height: 34px; width: 270px; font-size: 13px; font-weight: 600; text-align: center;']").click()
            driver.find_element_by_xpath("//span[text()='이번 달']").click()
            idx=idx+1
        elif idx==1:
            driver.find_element_by_xpath("//div[@style='background: rgb(255, 255, 255); color: rgb(51, 51, 51); line-height: 34px; width: 270px; font-size: 13px; font-weight: 600; text-align: center;']").click()
            driver.find_element_by_xpath("//span[text()='지난 달']").click()
            idx=idx+1
        tmp_dval=driver.find_element_by_xpath("//div[@style='background: rgb(255, 255, 255); color: rgb(51, 51, 51); line-height: 34px; width: 270px; font-size: 13px; font-weight: 600; text-align: center;']").get_attribute("innerHTML")
        tmp_date=tmp_dval[len(tmp_dval)-23:len(tmp_dval)-18]+"월"
        tmp_arr=[tmp_date]
        tmp_ddd=tmp_dval[len(tmp_dval)-20:len(tmp_dval)-17]
        loadin = get_text_by_xpath("//*[@class='highcharts-axis-labels highcharts-xaxis-labels']/*")[:3]
        while loadin != tmp_ddd:
            loadin = get_text_by_xpath("//*[@class='highcharts-axis-labels highcharts-xaxis-labels']/*")[:3]
        par = driver.find_elements_by_xpath("//th")
        for eachf in from_col:
            idxd=par.index(driver.find_element_by_xpath("//th/div/div/span[text()='"+eachf+"']/../../.."))
            tmp_arr.append(driver.find_element_by_xpath("//tr[@class='summary-row']/td["+str(idxd+1)+"]").get_attribute("innerHTML").replace(",","").replace("원",""))
        arr_pro.insert(0,tmp_arr)
        if driver.find_element_by_xpath("(//button[text()=' < '])").get_property('disabled'):
            break
        driver.find_element_by_xpath("(//button[text()=' < '])").click()
    df = pd.DataFrame(arr_pro)
    df.columns = arr_col
    df["광고비중"]="-"
    if store_id_exist:
        df2=get_pr_report("https://sell.smartstore.naver.com/#/bizadvisor/sales")
        df2=df2.copy()[["월간","결제금액"]]
        df2=df2.rename(columns = {"결제금액":"매출총액","수익률":"ROAS율"})
        df=pd.merge(df,df2)
        df["광고비중"]=(((df["전환액"].astype(float)*10000)/df["매출총액"].astype(float))//1/100).astype(str)+"%"
    df=df[["월간","노출수","클릭수","광고비","전환수","전환액","ROAS율","매출총액","광고비중"]]
    df[["노출수","클릭수","광고비","전환수","전환액","매출총액"]].replace(',','', regex=True).astype(int)
    return df

def get_pr_keyword(path):
    driver_execute(path)
    driver.find_element_by_xpath("//a[@data-nclicks-code='stu.selling']").click()
    driver.find_element_by_xpath('//div[@class="selectize-input items ng-valid ng-pristine has-options full has-items"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//div[@data-value="500"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//button[@class="btn btn-sm btn-default progress-button progress-button-dir-horizontal progress-button-style-top-line"]').click()
    time.sleep(1)
    tmp_file=find_file("Product",FILE_FOLDER)
    tmp_file=os.path.join(FILE_FOLDER,tmp_file)
    df=pd.read_csv(tmp_file, encoding='utf-8-sig')[["상품명","상품번호(스마트스토어)","할인가(PC)","대분류","중분류","소분류","세분류","대표이미지 URL"]]
    os.remove(tmp_file)
    df["카테고리명"]=""
    df.loc[(df['세분류'].isnull()),"카테고리명"]=df['대분류']+">"+df['중분류']+">"+df['소분류']
    df.loc[(df['세분류'].notnull()),"카테고리명"]=df['대분류']+">"+df['중분류']+">"+df['소분류']+">"+df['세분류']
    df2=df[["상품명","상품번호(스마트스토어)"]].copy()
    df2["추가홍보문구1"]=""
    df2["추가홍보문구2"]=""
    df2["제외키워드"]=""
    tmp_file2=find_file("※상품별 부가정보",FILE_FOLDER)
    if find_file(tmp_file2,FILE_FOLDER)!="False":
        tmp_file2=os.path.join(FILE_FOLDER,tmp_file2)
        df2=pd.read_csv(tmp_file2, encoding='utf-8-sig')
    df=pd.merge(df,df2,how="left")
    df2.to_csv(FILE_FOLDER + "/※상품별 부가정보.csv", encoding='utf-8-sig' ,index = False)
    return df

def get_ad_report(path):
    driver_execute(path)
    driver.find_element_by_xpath('//button[@class="btn-sm dropdown-toggle btn btn-default"]').click()
    driver.find_element_by_xpath("(//button[text()='필터 만들기'])").click()
    driver.find_element_by_xpath('//button[@style="word-break: keep-all;"]').click()
    driver.find_element_by_xpath('//div[@class="pl-2 filter-checkbox"][2]').click()
    driver.find_element_by_xpath('//button[@class="btn btn-sm btn-default-blue apply-button"]').click()
    driver.find_element_by_xpath("(//span[text()='다운로드'])").click()
    time.sleep(1)
    tmp_file=find_file("보고서(주1회)",FILE_FOLDER)
    tmp_file=os.path.join(FILE_FOLDER,tmp_file)
    df=pd.read_csv(tmp_file, encoding='utf-8-sig', skiprows=1)
    os.remove(tmp_file)
    return df

def get_ad_mass(path):
    mas_file=find_file("mas",FILE_FOLDER)
    if mas_file=="False":
        driver_execute(path)
        driver.find_element_by_xpath('//button[@class="btn btn-default dropdown-toggle"]').click()
        driver.find_element_by_xpath("(//button[text()='쇼핑검색'])").click()
        driver.find_element_by_xpath("(//button[@id='campaigns-dropdown'])").click()
        driver.find_element_by_xpath("(//div[@style='position: absolute; top: 0px; left: 0px; will-change: transform; transform: translate(0px, 32px);']/div[text()=' 스토어팜 ']/input)").click()
        driver.find_element_by_xpath("(//button[text()='다운로드'])").click()
        tmp=driver.find_element_by_xpath("(//tbody[@class='has-data ng-star-inserted']/tr[1]/td[11]/elena-mass-result)").get_attribute("innerHTML")
        while tmp=="<!---->":
            driver.find_element_by_xpath('//button[@class="btn btn-sm btn-default btn-icon"]').click()
            time.sleep(5)
            tmp=driver.find_element_by_xpath("(//tbody[@class='has-data ng-star-inserted']/tr[1]/td[11]/elena-mass-result)").get_attribute("innerHTML")
        driver.find_element_by_xpath("(//tbody[@class='has-data ng-star-inserted']/tr[1]/td[11]/elena-mass-result)").click()
        time.sleep(1)
        zp_fname=find_file("mas",FILE_FOLDER)
        os.chdir(FILE_FOLDER)
        zipfile.ZipFile(zp_fname).extractall()
        os.remove(zp_fname)
        mas_file=find_file("mas",FILE_FOLDER)
        mas_file=os.path.join(FILE_FOLDER,mas_file)
        df=pd.read_csv(mas_file, encoding='utf-8-sig', skiprows=1)
        os.remove(mas_file)
        df.to_csv(FILE_FOLDER+'/'+nowDate+'_mas.csv', encoding='utf-8-sig',index = False)
        mas_file=find_file("mas",FILE_FOLDER)
        mas_file=os.path.join(FILE_FOLDER,mas_file)
    else:
        mas_file=os.path.join(FILE_FOLDER,mas_file)
        df=pd.read_csv(mas_file, encoding='utf-8-sig')

    g_id_arr=df["광고그룹 ID"].copy().drop_duplicates().values
    for g_id in g_id_arr:
        path="https://manage.searchad.naver.com/customers/"+str(df_admin_dict["CUST_ID"])+"/adgroups/"+str(g_id)
        driver_execute(path)
        driver.find_element_by_xpath("//tr[@class='summary-row']")
        driver.find_element_by_xpath("//button[text()='다운로드']").click()
        each_mas=find_file("소재 목록 "+nowDate,FILE_FOLDER)
        while each_mas!="False":
            each_mas=os.path.join(FILE_FOLDER,each_mas)
        btn_listup=driver.find_element_by_xpath('(//button[@class="btn-sm btn-toggle dropdown-toggle btn btn-default"])')
        if not "행 표시: 200" in btn_listup.text:
            action = ActionChains(driver)
            action.move_to_element(driver.find_element_by_xpath('(//div[@class="footer-box"])')).perform()
            btn_listup.click()
            driver.find_element_by_xpath("(//button[text()=200])").click()
        each_mas=find_file("소재 목록 "+nowDate,FILE_FOLDER)
        while each_mas=="False":
            each_mas=find_file("소재 목록 "+nowDate,FILE_FOLDER)
        each_mas=os.path.join(FILE_FOLDER,each_mas)
        df3=pd.read_excel(each_mas, engine='openpyxl', skiprows=[1])
        os.remove(each_mas)
        df3["클릭률(%)"]=df3["클릭률(%)"].astype(str).str.replace(",","").astype(float)
        df3["전환율(%)"]=df3["전환율(%)"].astype(str).str.replace(",","").astype(float)
        df3["광고수익률(%)"]=df3["광고수익률(%)"].astype(str).str.replace(",","").astype(float)
        df3=df3.dropna(subset=['소재 ID'])
        for i in range(0,len(df3)):
            main_id=df3.iloc[i]["소재 ID"]
            status=df3.iloc[i]["상태"]
            if status=="중지 : 소재 연동제한":
                mas_file=find_file("mas",FILE_FOLDER)
                if mas_file!="False":
                    mas_file=os.path.join(FILE_FOLDER,mas_file)
                    os.remove(mas_file)
                remove_s_id(main_id)
            else:
                if status=="중지 : 소재 노출제한":
                    df.loc[df["소재 ID"]==main_id,"제한 사유"]=get_warn_why(main_id)
                tmp_img_url=driver.find_element_by_xpath("//td[@data-value='"+str(main_id)+"']/following-sibling::td/div/div/div/div").get_attribute("style")
                tmp_img_url=tmp_img_url.split('url("')[1].split('");')[0]
                df.loc[df["소재 ID"]==main_id,"대표이미지 URL"]=tmp_img_url
                for x in df3.columns:
                    df.loc[df["소재 ID"]==main_id,x]=df3.iloc[i][x]
    df.loc[df["노출상품명"]=="","노출상품명"]=df["기본상품명"]
    df=df.drop(columns=['쇼핑몰 상품ID', '소재 입찰가', '상품명', '카테고리'])
    df=df.drop(df.loc[df["상태"]=="중지 : 소재 연동제한"].index)
    return df

def get_sum_mass():
    global report_col_val
    global report_col_val_del
    report_col_val=['소재현황',   '노출상품명',  '광고그룹 이름',  '평균노출순위', '노출수',  '클릭수',  '전환수', '전환율(%)',  '총비용(VAT포함,원)', '전환매출액(원)',
    '소재 상태',    '입찰가',  '품질지수',   '소재 ID',  '광고그룹 ID', '쇼핑몰 상품 ID', '제한 사유']
    report_col_val_del=['\r\n',  '(',           ')\r\n',         '/',           '/',      '/',       '/',      '/',      '/',        '\r\n',
    '/',            "/",     '\r\n',         '/',        '/',         "\r\n",           ""]

    df_mas=df_web_dict["소재현황리스트"]
    df_warn=df_web_dict["☆점검소재리스트"]
    df_mas_sum=df_mas.copy().groupby(['기본상품명','쇼핑몰 상품 ID','네이버쇼핑 카테고리','대표이미지 URL','상품가격']).sum().reset_index()[[
        '기본상품명','쇼핑몰 상품 ID','네이버쇼핑 카테고리','대표이미지 URL','상품가격',
        '노출수','클릭수','전환수','총비용(VAT포함,원)','전환매출액(원)'
    ]]
    tmp_file2=find_file("※상품별 부가정보",FILE_FOLDER)
    if find_file(tmp_file2,FILE_FOLDER)!="False":
        tmp_file2=os.path.join(FILE_FOLDER,tmp_file2)
        df2=pd.read_csv(tmp_file2, encoding='utf-8-sig')
        df2.rename(columns = {'상품번호(스마트스토어)' : '쇼핑몰 상품 ID', '상품명':'기본상품명'}, inplace = True)
        df_mas_sum=pd.merge(df_mas_sum,df2)
    df_mas_sum=df_mas_sum.sort_values(by=['노출수','클릭수','총비용(VAT포함,원)'], ascending=[False,False,True])
    df_mas_sum["클릭률"]=(df_mas_sum["클릭수"]*10000/df_mas_sum["노출수"]//1)/100
    df_mas_sum["전환율"]=(df_mas_sum["전환수"]*10000/df_mas_sum["클릭수"]//1)/100
    df_mas_mean=df_mas.copy()[['쇼핑몰 상품 ID','평균노출순위','입찰가']].groupby('쇼핑몰 상품 ID').mean().reset_index()
    df_mas_mean['평균노출순위']=df_mas_mean['평균노출순위']*10//1/10
    df_mas_mean['입찰가']=df_mas_mean['입찰가']//1
    df_mas_sum=pd.merge(df_mas_sum,df_mas_mean,on="쇼핑몰 상품 ID",how="left")
    df_warn=df_warn.copy().drop_duplicates(subset=['쇼핑몰 상품 ID'])[['쇼핑몰 상품 ID','제한 사유']]
    df_mas_sum=pd.merge(df_mas_sum,df_warn,on="쇼핑몰 상품 ID",how="left")
    df_mas_sum=df_mas_sum[['제외키워드','대표이미지 URL','기본상품명','상품가격','네이버쇼핑 카테고리',
    '노출수','클릭수','클릭률','전환수','전환율','총비용(VAT포함,원)','전환매출액(원)'
    ,'쇼핑몰 상품 ID','추가홍보문구1','추가홍보문구2','평균노출순위','입찰가','제한 사유']]
    df_grp_arr=df_mas['광고그룹 이름'].copy().drop_duplicates().values
    df_pid_arr=df_mas['쇼핑몰 상품 ID'].copy().drop_duplicates().values
    for df_idx in range(0,len(df_mas_sum)):
        this_pro_id=df_mas_sum.iloc[df_idx]['쇼핑몰 상품 ID']
        tmp_df=df_mas[df_mas['쇼핑몰 상품 ID'].astype(str)==str(this_pro_id)].copy()
        this_pro_gid=df_mas[df_mas['쇼핑몰 상품 ID'].astype(str)==str(this_pro_id)].copy()
        for sub_idx in range(0,len(tmp_df)):
            if not "소재"+str(sub_idx+1) in df_mas_sum.columns:
                df_mas_sum["소재"+str(sub_idx+1)]=""
            tmp_df_val=""
            v_idx=0
            for each_col in report_col_val:
                tmp_df_val = tmp_df_val + str(tmp_df.iloc[sub_idx][each_col]).replace(".0","")
                tmp_df_val = tmp_df_val + report_col_val_del[v_idx]
                v_idx=v_idx+1
            df_mas_sum["소재"+str(sub_idx+1)].iat[df_idx]=str(tmp_df_val)
    return df_mas_sum

def get_status_mass():
    global wait_val
    global goodgg_val
    global badgg_val
    global newgg_val
    global warngg_val
    global gg_val
    global gg_col

    # 노출 상태 종류
    goodgg_val='양호 소재'
    goodgg_col='FF00B050'
    badgg_val='불량 소재'
    badgg_col='FFFF0000'
    chapriup_val='입찰가 변경인상'
    chapriup_col='FF0070C0'
    chapridwn_val='입찰가 변경인하'
    chapridwn_col='FFF4B084'
    wait_val='대기 소재'
    wait_col='FFC65911'
    newgg_val='신규 소재'
    newgg_col='FFFFFF00'
    offgg_val='off 소재'
    offgg_col='FFAEAAAA'
    warngg_val='소재 점검 필요'
    warngg_col='FFFFC000'
    emptygg_val='빈 그룹'
    emptygg_col='FFFFFFFF'

    gg_val=[goodgg_val,badgg_val,chapriup_val,chapridwn_val,wait_val,newgg_val,offgg_val,warngg_val,emptygg_val]
    gg_col=[goodgg_col,badgg_col,chapriup_col,chapridwn_col,wait_col,newgg_col,offgg_col,warngg_col,emptygg_col]

    df_status=df_web_dict["광고소재리스트"].copy()
    df_report=df_web_dict["스토어팜-보고서"].copy()

    cut_rank=df_admin_dict['삭제순위(이하)']
    cut_show=df_admin_dict['삭제노출(이하)']
    cut_clk=df_admin_dict['삭제클릭(이하)']
    cut_buy=df_admin_dict['삭제전환(이하)']
    cut_pri_up=df_admin_dict['삭제입찰가(이상)']
    cut_pri_dwn=df_admin_dict['삭제입찰가(이하)']

    price_dwn_cut=df_admin_dict['입찰가상승순위(이상)']
    price_dwn_val=df_admin_dict['입찰가상승값']
    price_up_cut=df_admin_dict['입찰가하강순위(이하)']
    price_up_val=df_admin_dict['입찰가하강값']

    for i in range(0,len(df_report)):
        s_id=df_report.iloc[i]["소재"]
        for x in df_report.columns:
            df_status.loc[df_status["소재 ID"]==s_id,x]=df_report.iloc[i][x]
    df_status=df_status.drop(columns=['캠페인유형','캠페인','소재','광고그룹'])
    df_status.loc[df_status['노출상품명'].isnull(),"노출상품명"]=df_status['기본상품명']
    df_status_col=['평균노출순위','노출수','클릭수','전환수','총비용(VAT포함,원)','입찰가','전환매출액(원)']
    for each_col in df_status_col:
        df_status[each_col]=df_status[each_col].astype(str).str.replace(",","").astype(float)

    df_status["소재현황"]=goodgg_val
    tmp_df=df_status.loc[(df_status['평균노출순위']>=price_dwn_cut)].copy()
    df_status.loc[(df_status['평균노출순위']>=price_dwn_cut),"소재현황"]=chapriup_val + df_status["입찰가"].astype(int).astype(str) + ">" + (df_status["입찰가"].astype(int)+int(price_dwn_val)*((df_status['평균노출순위'].astype(float)-float(price_dwn_cut))//10+1)).astype(int,errors='ignore').astype(str)
    df_status.loc[(df_status['평균노출순위']>=price_dwn_cut),"입찰가"]=(df_status["입찰가"].astype(int)+int(price_dwn_val)*((df_status['평균노출순위'].astype(float)-float(price_dwn_cut))//10+1)).astype(int,errors='ignore')
    df_status.loc[(df_status['평균노출순위']<=price_up_cut),"소재현황"]=chapridwn_val + df_status["입찰가"].astype(int).astype(str) + ">" + (df_status["입찰가"].astype(int)+int(price_up_val)*((float(price_up_cut)-df_status['평균노출순위'].astype(float))//10+1)).astype(int,errors='ignore').astype(str)
    df_status.loc[(df_status['평균노출순위']<=price_up_cut),"입찰가"]=(df_status["입찰가"].astype(int)+int(price_up_val)*((float(price_up_cut)-df_status['평균노출순위'].astype(float))//10+1)).astype(int,errors='ignore')
    df_status.loc[(df_status['평균노출순위']<=cut_rank) & (df_status['노출수']<=cut_show) & (df_status['클릭수']<=cut_clk) & (df_status['전환수']<=cut_buy)
    & ~((cut_pri_dwn < df_status['입찰가']) & (df_status['입찰가'] < cut_pri_up)),"소재현황"]=badgg_val
    if nextWeekDay==df_admin_dict["정기보고(요일)"]:
        df_status.loc[(df_status['노출수']==0),"소재현황"]=badgg_val
    else:
        df_status.loc[(df_status['노출수']==0),"소재현황"]=wait_val
    df_status.loc[df_status['제한 사유'].notnull(),"소재현황"]=warngg_val

    df_pid_arr=df_status['쇼핑몰 상품 ID'].copy().drop_duplicates().values
    df_grp_arr=df_status['광고그룹 이름'].copy().drop_duplicates().values
    for df_pid in df_pid_arr:
        tmp_df=df_status[df_status['쇼핑몰 상품 ID']==df_pid]
        df_grp_arr_each=tmp_df['광고그룹 이름'].copy().drop_duplicates().values
        grp_minus=np.setdiff1d(df_grp_arr,df_grp_arr_each)
        for x in grp_minus:
            t_idx=0
            next_each=len(df_status)
            df_status.loc[next_each,"소재현황"]=emptygg_val
            df_status.loc[next_each,"기본상품명"]=tmp_df["기본상품명"].iloc[0]
            tmp_new_key=get_rel_key(tmp_df["노출상품명"].iloc[t_idx],df_pid)
            while tmp_new_key in df_status["노출상품명"].values.tolist():
                tmp_new_key=get_rel_key(tmp_df["노출상품명"].iloc[t_idx],df_pid)
                tmp_df=df_status[df_status['쇼핑몰 상품 ID']==df_pid]
                if t_idx<len(tmp_df)-1:
                    t_idx=t_idx+1
                else:
                    tmp_new_key="-"
                    break
            df_status.loc[next_each,"노출상품명"]=tmp_new_key
            df_status.loc[next_each,"대표이미지 URL"]=tmp_df["대표이미지 URL"].iloc[0]
            df_status.loc[next_each,"상품가격"]=tmp_df["상품가격"].iloc[0]
            df_status.loc[next_each,"네이버쇼핑 카테고리"]=tmp_df["네이버쇼핑 카테고리"].iloc[0]
            df_status.loc[next_each,"네이버쇼핑 상품 ID"]=tmp_df["네이버쇼핑 상품 ID"].iloc[0]
            df_status.loc[next_each,"광고그룹 이름"]=x
            df_status.loc[next_each,"광고그룹 ID"]=df_status[df_status['광고그룹 이름']==x]['광고그룹 ID'].iloc[0]
            df_status.loc[next_each,"쇼핑몰 상품 ID"]=df_pid
    df_status=df_status.sort_values(by=['쇼핑몰 상품 ID','노출수','클릭수','총비용(VAT포함,원)','소재 상태','광고그룹 이름'], ascending=[False,False,False,True,False,True])
    for df_pid in df_pid_arr:
        tmp_df=df_status[df_status["쇼핑몰 상품 ID"]==df_pid]
        if tmp_df["소재현황"].iloc[0]!=warngg_val:
            st_x=10+len(tmp_df[tmp_df["소재현황"]==badgg_val])
            tmp_df=df_status[df_status["쇼핑몰 상품 ID"]==df_pid].iloc[st_x:]["광고그룹 이름"]
            df_status.loc[(df_status["쇼핑몰 상품 ID"]==df_pid) & (df_status["광고그룹 이름"].isin(tmp_df.values))
            & (df_status["소재 상태"]=="off"),"소재현황"]=offgg_val
            tmp_df=df_status[df_status["쇼핑몰 상품 ID"]==df_pid].iloc[:st_x]["광고그룹 이름"]
            df_status.loc[(df_status["쇼핑몰 상품 ID"]==df_pid) & (df_status["광고그룹 이름"].isin(tmp_df.values))
            & ((df_status["소재현황"]==emptygg_val) | (df_status["소재현황"]==offgg_val)),"소재현황"]=newgg_val
    return df_status

def get_warn_mass():
    df=df_web_dict["광고소재리스트"]
    df_warn=df[df["상태"]=="중지 : 소재 노출제한"]
    for i in range(0,len(df_warn)):
        s_id=df_warn["소재 ID"].iloc[i]
        s_name=df_warn["기본상품명"].iloc[i]
        img_link=df_warn["대표이미지 URL"].iloc[i]
        tmp_img_link=img_link.split(".")
        urllib.request.urlretrieve(img_link, OUTPUT_FOLDER+"\\"+s_name+"_"+s_id+"."+tmp_img_link[len(tmp_img_link)-1])
    return df_warn

def remove_s_id(s_id):
    if driver.find_element_by_xpath('(//button[@class="btn-sm btn-toggle dropdown-toggle btn btn-default"])').text!="행 표시: 200":
        driver.find_element_by_xpath('//button[@class="btn-sm btn-toggle dropdown-toggle btn btn-default"]').click()
        driver.find_element_by_xpath("(//button[text()=200])").click()
    # 행표시 200개 / 다음페이지 이동 / 노출가능 클릭
    action = ActionChains(driver)
    action.move_to_element(driver.find_element_by_xpath('//div[@class="inner-left"]')).perform()
    driver.find_element_by_xpath('//td[@data-value="'+s_id+'"]/preceding-sibling::td[1]/preceding-sibling::td[1]/preceding-sibling::td[1]/input').click()
    action.move_to_element(driver.find_element_by_xpath('//div[@class="inner-left"]')).perform()
    driver.find_element_by_xpath('//button[@class="ml-2 btn btn-default btn-sm"]').click()
    driver.find_element_by_xpath('//button[@class="btn btn-primary Confirm_modal-button__1Kwk6 btn btn-secondary"]').click()

def get_warn_why(s_id):
    if driver.find_element_by_xpath('(//button[@class="btn-sm btn-toggle dropdown-toggle btn btn-default"])').text!="행 표시: 200":
        driver.find_element_by_xpath('//button[@class="btn-sm btn-toggle dropdown-toggle btn btn-default"]').click()
        driver.find_element_by_xpath("(//button[text()=200])").click()
    # 행표시 200개 / 다음페이지 이동 / 노출가능 클릭
    row_posi = driver.find_element_by_xpath('//td[@data-value="'+s_id+'"]/preceding-sibling::td[1]/span/a')
    action = ActionChains(driver)
    try:
        action.move_to_element(row_posi).perform()
        row_posi.click()
    except Exception as e:
        row_posi_parent=driver.find_element_by_xpath('//td[@data-value="'+s_id+'"]/parent::tr/preceding-sibling::tr[1]/td[1]')
        action.move_to_element(row_posi_parent).perform()
        row_posi.click()
    while driver.find_element_by_xpath('//div[@class="list-dot"]').text=="":
        print(driver.find_element_by_xpath('//div[@class="list-dot"]').text)
    row_detail=[]
    for row_detail_each in driver.find_elements_by_xpath('//div[@class="list-dot"]'):
        row_detail.append(row_detail_each.text.replace("더 알아보기",""))
    driver.find_element_by_xpath('//button[@class="modal-button btn btn-default"]').click()
    return "\r\n".join(row_detail)

def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return str(string)

# -------------------------------------------------------------------------------

def init():
    global THIS_FOLDER
    global nowDate
    global nowWeekDay
    global nextWeekDay
    global isFirstWeek
    global df_admin
    global driver_exist
    global login_status
    global weekDel
    global weekHold

    weekDel="☆"
    weekHold="★"

    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    driver_exist=False
    login_status=False
    days=["월","화","수","목","금","토","일"]
    nowDate=datetime.datetime.now().strftime('%Y%m%d')
    if int(datetime.datetime.now().strftime('%d'))<8:
        isFirstWeek=True
    else:
        isFirstWeek=False
    nowWeekDay=days[datetime.datetime.today().weekday()]
    nextWeekDay=days[(datetime.datetime.today().weekday()+8)%7]

def load_df_admin():
    global df_admin_col
    # df_admin_col=["스토어","CUST_ID","광고","캠페인","CAMP_ID","정기보고(요일)","보고수신메일","최종보고일",
    # "스마트스토어 ID","스마트스토어 PW","검색광고 ID","검색광고 PW","검색어보고서","스토어팜보고서",
    # "키워드 상위(%)","키워드 최대갯수",
    # "삭제순위(이하)","삭제노출(이하)","삭제클릭(이하)","삭제전환(이하)","삭제입찰가(이상)","삭제입찰가(이하)",
    # "입찰가상승순위(이상)","입찰가상승값","입찰가하강순위(이하)","입찰가하강값","기본입찰가"]
    file_admin=find_file("※광고주관리",THIS_FOLDER)
    if file_admin!="False":
        admin_file=os.path.join(THIS_FOLDER,file_admin)
        df_admin=pd.read_csv(file_admin, encoding='utf-8-sig')
        df_admin_col=df_admin.columns
    else:
        df_admin=pd.DataFrame(columns=df_admin_col)
        df_admin.to_csv(THIS_FOLDER+"/※광고주관리.csv", encoding='utf-8-sig', index=False)
    return df_admin

def update(admin_df):
    global FILE_FOLDER
    global OUTPUT_FOLDER
    global df_admin_dict

    df_admin_dict={}
    for key in df_admin_col:
        df_admin_dict[key]=admin_df[key]
    filedel_list=os.listdir(THIS_FOLDER)
    for filedel in filedel_list:
        if not "※" in filedel and weekDel in filedel and nowWeekDay=="일":
            del_fname=os.path.join(THIS_FOLDER,filedel)
            os.remove(del_fname)
        elif not "※" in filedel and weekHold in filedel and nowWeekDay!="일":
            del_fname=os.path.join(THIS_FOLDER,filedel)
            os.rename(del_fname, del_fname.replace(weekHold,weekDel))
    FILE_FOLDER=os.path.join(THIS_FOLDER,df_admin_dict["스토어"])
    createFolder(FILE_FOLDER)
    filedel_list=os.listdir(FILE_FOLDER)
    for filedel in filedel_list:
        if ".csv" in filedel:
            if not "※" in filedel and not nowDate in filedel:
                del_fname=os.path.join(FILE_FOLDER,filedel)
                os.remove(del_fname)
        elif df_admin_dict['광고'] in filedel:
            del_foldname=os.path.join(FILE_FOLDER,filedel)
            shutil.rmtree(del_foldname)
    OUTPUT_FOLDER = FILE_FOLDER+'//'+nowDate+"_"+df_admin_dict['광고']
    createFolder(OUTPUT_FOLDER)

def get_data_file():
    global df_web_dict
    global c_id

    df_web_col=["상품리스트",
    "광고대비전환","광고소재리스트","인기검색어",
    "검색어-보고서","스토어팜-보고서"]
    df_web_dict={}
    for key in df_web_col:
        key_file=nowDate+"_"+key+".csv"
        key_file=os.path.join(FILE_FOLDER,key_file)
        if find_file(key,FILE_FOLDER)!="False":
            df_web_dict[key]=pd.read_csv(key_file, encoding='utf-8-sig')
        elif store_id_exist and key=="상품리스트":
            df_web_dict[key]=get_pr_keyword("https://sell.smartstore.naver.com/#/products/origin-list")
            df_web_dict[key].to_csv(key_file, encoding='utf-8-sig',index = False)
        else:
            if key=="광고대비전환":
                df_web_dict[key]=get_adp_report("https://manage.searchad.naver.com/customers/"+str(df_admin_dict["CUST_ID"])+"/campaigns/"+str(df_admin_dict["CAMP_ID"]))
            elif key=="광고소재리스트":
                df_web_dict[key]=get_ad_mass("https://manage.searchad.naver.com/customers/"+ str(df_admin_dict["CUST_ID"])+"/tool/mass")
            elif key=="인기검색어":
                df_web_dict[key]=get_cate_key("https://datalab.naver.com/shoppingInsight/sCategory.naver")
            elif key=="검색어-보고서":
                df_web_dict[key]=get_ad_report("https://manage.searchad.naver.com/customers/"+ str(df_admin_dict["CUST_ID"])+"/reports/"+ str(df_admin_dict["검색어보고서"]))
            elif key=="스토어팜-보고서":
                df_web_dict[key]=get_ad_report("https://manage.searchad.naver.com/customers/"+ str(df_admin_dict["CUST_ID"])+"/reports/"+ str(df_admin_dict["스토어팜보고서"]))
            df_web_dict[key].to_csv(key_file, encoding='utf-8-sig',index = False)

def insert_row(row_number, df, row_value):
    start_upper = 0
    end_upper = row_number
    start_lower = row_number
    end_lower = df.shape[0]
    upper_half = [*range(start_upper, end_upper, 1)]
    lower_half = [*range(start_lower, end_lower, 1)]
    lower_half = [x.__add__(1) for x in lower_half]
    index_ = upper_half + lower_half
    df.index = index_
    df.loc[row_number] = row_value
    df = df.sort_index()
    return df

def get_new_sid():
    df_mas=df_web_dict["소재현황리스트"].copy()
    df_mas=df_mas[(df_mas["소재현황"]==newgg_val) & (df_mas["소재 ID"].isnull()) & (df_mas["노출상품명"].notnull() & df_mas["노출상품명"]!="-")]
    df_mas=df_mas[["광고그룹 ID","네이버쇼핑 상품 ID","입찰가","노출상품명"]]
    if len(df_mas)==0:
        return
    if store_id_exist:
        df_sub=df_web_dict["상품리스트"][["상품번호(스마트스토어)","추가홍보문구1","추가홍보문구2"]]
        df_mas=pd.merge(df_mas,df_sub,left_on="네이버쇼핑 상품 ID",right_on="상품번호(스마트스토어)",how="left")
        df_mas=df_mas.drop(columns=['상품번호(스마트스토어)'])
    for i in range(0,5):
        df_mas=insert_row(0,df_mas,df_mas.columns)
    df_mas["입찰가"]=int(df_admin_dict["기본입찰가"])
    df_mas["네이버쇼핑 상품 ID"]=str(df_admin_dict["네이버쇼핑 상품 ID"])
    return df_mas
def get_on_sid():
    df_mas=df_web_dict["소재현황리스트"].copy()
    df_mas=df_mas.loc[((df_mas["소재현황"]==newgg_val) | ((df_mas["소재현황"]==goodgg_val) & (df_mas["소재 상태"]=="off")))
    & (df_mas["소재 ID"].notnull()) ][["광고그룹 ID","소재 ID","소재 상태"]]
    if len(df_mas)==0:
        return
    df_mas["소재 상태"]="on"
    for i in range(0,5):
        df_mas=insert_row(0,df_mas,df_mas.columns.tolist())
    return df_mas
def get_chgpri_sid():
    df_mas=df_web_dict["소재현황리스트"].copy()
    df_mas=df_mas[df_mas["소재현황"].str.contains("입찰가")][["광고그룹 ID","소재 ID","입찰가"]]
    if len(df_mas)==0:
        return
    for i in range(0,5):
        df_mas=insert_row(0,df_mas,df_mas.iloc[0])
    df_mas["입찰가"]=df_mas["입찰가"].astype(int)
    return df_mas
def get_del_sid():
    df_mas=df_web_dict["소재현황리스트"].copy()
    df_mas=df_mas.loc[((df_mas["소재현황"]==badgg_val) | ((df_mas["제한 사유"].notnull()) & (df_mas["소재 상태"]=="off")))][["광고그룹 이름","광고그룹 ID","소재 ID"]]
    if len(df_mas)==0:
        return
    return df_mas

def make_import_file():
    df_import_col=["☆점검소재리스트","소재현황리스트","소재부분합리스트",
    "☆신규소재리스트","☆소재상태on리스트","☆입찰가변경리스트","☆삭제소재리스트"]
    for key in df_import_col:
        key_file=nowDate+"_"+key+".csv"
        key_file=os.path.join(OUTPUT_FOLDER,key_file)
        if key=="☆점검소재리스트":
            df_web_dict[key]=get_warn_mass()
        elif key=="소재현황리스트":
            df_web_dict[key]=get_status_mass()
        elif key=="소재부분합리스트":
            df_web_dict[key]=get_sum_mass()
        elif key=="☆신규소재리스트":
            df_web_dict[key]=get_new_sid()
        elif key=="☆소재상태on리스트":
            df_web_dict[key]=get_on_sid()
        elif key=="☆입찰가변경리스트":
            df_web_dict[key]=get_chgpri_sid()
        elif key=="☆삭제소재리스트":
            df_web_dict[key]=get_del_sid()
        if isinstance(df_web_dict[key], pd.DataFrame):
            df_web_dict[key].to_csv(key_file, encoding='utf-8-sig',index = False)
            if str(df_admin_dict["작업일자"])!=str(nowDate):
                import_data_to_ad(key,key_file)

def ready_to_ad(txt,key_file):
    path="https://manage.searchad.naver.com/customers/"+str(df_admin_dict["CUST_ID"])+"/mass"
    driver_execute(path)
    if driver.find_element_by_xpath("(//button[@class='dropdown-toggle btn btn-default'])").text!=df_admin_dict["광고"]:
        driver.find_element_by_xpath("(//button[@class='dropdown-toggle btn btn-default'])").click()
        driver.find_element_by_xpath("(//li[text()='"+df_admin_dict["광고"]+"'])").click()
        driver.find_element_by_xpath("(//a[text()='대량 등록/수정'])").click()
    driver.find_element_by_xpath("(//span[text()='선택하세요'])").click()
    driver.find_element_by_xpath("(//button[text()='"+txt+"'])").click()
    driver.find_element_by_css_selector("input[type='file']").send_keys(key_file)
    driver.find_element_by_xpath("(//button[text()='등록하기'])").click()

def import_data_to_ad(key,key_file):
    df=df_web_dict[key]
    if key=="☆신규소재리스트":
        ready_to_ad("쇼핑몰 상품 소재 등록",key_file)
    elif key=="☆소재상태on리스트":
        ready_to_ad("소재 On-Off 수정",key_file)
    elif key=="☆입찰가변경리스트":
        ready_to_ad("소재 입찰가 수정",key_file)
    elif key=="☆삭제소재리스트":
        g_id_arr=df["광고그룹 ID"].copy().drop_duplicates().values
        for g_id in g_id_arr:
            path="https://manage.searchad.naver.com/customers/"+str(df_admin_dict["CUST_ID"])+"/adgroups/"+str(g_id)
            driver_execute(path)
            s_id_arr=df.loc[df["광고그룹 ID"]==g_id,"소재 ID"].copy().drop_duplicates().values
            # for s_id in s_id_arr:
            #     remove_s_id(s_id)
    # elif key=="☆점검소재리스트":
    #     s_id_arr=df.copy().drop_duplicates().values
    #     for s_id in s_id_arr:
    #         path="https://manage.searchad.naver.com/customers/"+str(df_admin_dict["CUST_ID"])+"/ads/"+str(s_id)
    #         driver_execute(path)
    #         if df.loc[df["소재 ID"]==s_id]["제한 사유"].contains("반복나열"):
    #             change_name_s_id(s_id)
    #         elif df.loc[df["소재 ID"]==s_id]["제한 사유"].contains("이미지 내 텍스트"):
    #             change_img_s_id(s_id)

def change_name_s_id(s_id):
    driver.find_element_by_xpath("(//button[text()='수정'])").click()
    tmp_name=driver.find_element_by_xpath("(//input[@class='form-control 0 form-control'])").ad_text
    if tmp_name=="":
        tmp_name=""
    driver.find_element_by_xpath("(//input[@class='form-control 0 form-control'])").send_keys(trim_name(tmp_key))
    print("chg_sid")
def change_img_s_id(s_id):
    print("img_sid")

def make_report_file():
    global wb
    global ws_desc
    global ws_chg
    global ws_ad
    global ws_key

    wb = load_workbook(THIS_FOLDER+'\※마케팅보고.xlsx')
    ws_desc = wb['간이보고']
    ws_chg = wb['월간전환상세']
    ws_ad = wb['주간광고상세']
    ws_key = wb['주간키워드상세']

    set_desc_report()
    set_chg_report()
    set_ad_report()
    set_key_report()

    wb.save(OUTPUT_FOLDER+'/'+nowDate+'_☆마케팅보고서.xlsx')

def set_data_to_ad():
    driver_logout()
    if str(df_admin_dict["작업일자"])!=str(nowDate):
        set_df_admin(df_admin_dict["CAMP_ID"],"작업일자",nowDate)
        set_df_admin(df_admin_dict["CAMP_ID"],"광고잔액",df_admin_dict["광고잔액"])
        set_df_admin(df_admin_dict["CAMP_ID"],"입찰가하강순위(이하)",price_up_cut)
        set_df_admin(df_admin_dict["CAMP_ID"],"입찰가상승순위(이상)",price_dwn_cut)

if __name__ == '__main__':

    conti=input("if you dont want to test, e")
    if conti!="e":
        global THIS_FOLDER
        global FILE_FOLDER
        global nowWeekDay
        global weekDel
        THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
        FILE_FOLDER = "C:/Users/dltjd/ALINE/mysite/ex/알리네 간판"
        nowWeekDay="화"
        weekDel="☆"
        file_admin=find_file("상품리스트","C:/Users/dltjd/ALINE/mysite/ex/알리네 간판")
        if file_admin!="False":
            admin_file=os.path.join("C:/Users/dltjd/ALINE/mysite/ex/알리네 간판",file_admin)
            df_admin=pd.read_csv(admin_file, encoding='utf-8-sig')
        while True:
            pid=input("pid")
            pname=df_admin.loc[df_admin["상품번호(스마트스토어)"].astype(str)==str(pid)]["상품명"].values[0]
            print(pname)
            print(get_rel_key(pname,pid))

    timer_start()
    init()

    global store_id_exist
    df_admin=load_df_admin()
    for df_idx in range(0,len(df_admin)):
        update(df_admin.iloc[df_idx])
        if df_admin_dict["스마트스토어 ID"]!="":
            store_id_exist=True
        else:
            store_id_exist=False
        if df_admin_dict["광고"]=="쇼핑검색":
            get_data_file()
            make_import_file()
            make_report_file()
            set_data_to_ad()

    if driver_exist:
        driver.close()
    timer_chk()
