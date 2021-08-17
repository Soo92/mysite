import csv
import datetime
import openpyxl
import clipboard
import requests
import hashlib, hmac, base64, requests, time, os
import urllib.request, json
from urllib.request import urlopen, Request
import random
import shutil
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.styles import Border, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import zipfile
import os, time, numpy as np, pandas as pd
import xlsxwriter
from bs4 import BeautifulSoup    #BeautifulSoup import
import time

def find_file(fname):
    file_flist=os.listdir(FILE_FOLDER)
    for file_f in file_flist:
        if file_f.find(fname)>-1:
            return file_f
    return "False"

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)

def test():
    while True:
        clk=input("c or e or any~~")
        try:
            print(clk)
            if clk=="c":
                key=input("key")
                driver.find_element_by_xpath(key).click()
            elif clk=="e":
                break
            elif clk=="m":
                key=input("key")
                posi_m=driver.find_element_by_xpath(key)
                action = ActionChains(driver)
                action.move_to_element(posi_m).perform()
            else:
                key=input("key")
                print(driver.find_element_by_xpath(key).get_attribute("innerHTML"))
        except Exception as e:
            print("Error:" + str(e))

def generate(timestamp, method, uri, secret_key):
    message = "{}.{}.{}".format(timestamp, method, uri)
#     hash = hmac.new(bytes(secret_key, "utf-8"), bytes(message, "utf-8"), hashlib.sha256)
    hash = hmac.new(secret_key.encode("utf-8"), message.encode("utf-8"), hashlib.sha256)
    hash.hexdigest()
    return base64.b64encode(hash.digest())

def get_header(method, uri, api_key, secret_key, customer_id):
    timestamp = str(int(time.time() * 1000))
    signature = generate(timestamp, method, uri, SECRET_KEY)
    return {'Content-Type': 'application/json; charset=UTF-8', 'X-Timestamp': timestamp, 'X-API-KEY': API_KEY, 'X-Customer': str(CUSTOMER_ID), 'X-Signature': signature}

def call_RelKwd(_kwds_string):
    global BASE_URL,CUSTOMER_ID,API_KEY,SECRET_KEY
    BASE_URL = 'https://api.naver.com'
    CUSTOMER_ID = '392590'
    API_KEY = '01000000008fa2a584355277148cf5b1792f0a1650becf66b049812186ce69dbfb7cbf4ec3'
    SECRET_KEY = 'AQAAAACPoqWENVJ3FIz1sXkvChZQySFY24NP0GvvyT3R14cXaQ=='

    uri = '/keywordstool'
    method = 'GET'
    prm = {'hintKeywords' : _kwds_string , 'showDetail':1}
    # ManageCustomerLink Usage Sample
    returnData = None
    df = pd.DataFrame()
    # print(_kwds_string)
    repeat_time=0.5
    wait_time=3
    while returnData is None:
        try:
            r = requests.get(BASE_URL + uri, params=prm, headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
            returnData = r.json()
            df = pd.DataFrame(returnData['keywordList'])
        except Exception as e:
            if 'code' in returnData:
                time.sleep(wait_time)
                wait_time=wait_time+1
                print(_kwds_string)
                print(wait_time)
            time.sleep(repeat_time)
            r = requests.get(BASE_URL + uri, params=prm, headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
            returnData = r.json()
            df = pd.DataFrame(returnData['keywordList'])
            pass
        if not 'keywordList' in returnData:
            print(_kwds_string)
            print(returnData)
            return False

    # df.to_csv(FILE_FOLDER+"/sdfasdf.csv", encoding='utf-8-sig')
    df['총검색']=df['monthlyMobileQcCnt'].astype(str).replace("< 10",0).astype(float)+df['monthlyPcQcCnt'].astype(str).replace("< 10",0).astype(float)
    df['총클릭']=df['monthlyAveMobileClkCnt'].astype(float)+df['monthlyAvePcClkCnt'].astype(float)
    df.rename({'compIdx':'경쟁도',       'monthlyAveMobileClkCnt':'평균클릭(폰)',       'monthlyAveMobileCtr':'평균클릭률(폰)',
       'monthlyAvePcClkCnt':'평균클릭(PC)',       'monthlyAvePcCtr':'평클릭률(PC)',       'monthlyMobileQcCnt':'검색(폰)',
       'monthlyPcQcCnt': '검색(PC)',       'plAvgDepth':'노출광고수',       'relKeyword':'연관키워드'},axis=1,inplace=True)
    rate=0.95
    if len(df)>1:
        tmp_df=df[(df['총검색'] > df['총검색'].quantile(rate)) & (df['총클릭'] > df['총클릭'].quantile(rate))]
        while len(tmp_df)==0:
            tmp_df=df[(df['총클릭'] > df['총클릭'].quantile(rate))]
            rate=rate-0.05
        df=tmp_df
        df=df.sort_values(by="총클릭", ascending=False)
    return df

def make_new_name(old_name,df_rel):
    key_arr=old_name.split(" ")
    new_arr=[]
    for tmp_key in key_arr:
        tmp_rel=df_rel[df_rel["노출키워드"]==tmp_key]
        tmp_rel=tmp_rel["연관키워드"].copy().drop_duplicates().values
        if not tmp_key in tmp_rel:
            tmp_rel=np.append(tmp_rel,tmp_key)
        if len(tmp_rel)>0:
            rel_idx=random.randint(0, len(tmp_rel)-1)
            new_arr.append(tmp_rel[rel_idx])
    return " ".join(new_arr)

def get_cate_key(cate,arr_pro):
    path = 'https://datalab.naver.com/shoppingInsight/sCategory.naver'
    driver_execute(path)
    cate_arr=cate.split(">>")
    # 기기별 전체 선택
    driver.find_element_by_xpath('//*[@id="18_device_0"]').click()
    # 성별 전체 선택
    driver.find_element_by_xpath('//*[@id="19_gender_0"]').click()
    # 연령별 전체 선택
    driver.find_element_by_xpath('//*[@id="20_age_0"]').click()
    # 분류 & 기간 선택
    for i in range(0, len(cate_arr)):
        driver.find_element_by_xpath("(//span[contains(@class,'select_btn')])["+str(i+1)+"]").click()
        driver.find_element_by_xpath("(//a[text()='"+cate_arr[i]+"'])").click()
    # 조회하기 클릭
    driver.find_element_by_xpath('//*[@id="content"]/div[2]/div/div[1]/div/a').click()
    time.sleep(1)
    for p in range(0, 25):
        # 인기검색어 가져오기
        for i in range(1, 21):
            keyword_path = '//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[1]/ul/li[{}]/a'.format(i)
            key_num=driver.find_element_by_xpath(keyword_path).text.split("\n")[0]
            key_word=driver.find_element_by_xpath(keyword_path).text.split("\n")[1].replace(" ","")
            while(int(key_num)!=p*20+i):
                time.sleep(0.1)
                keyword_path = '//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[1]/ul/li[{}]/a'.format(i)
                key_num=driver.find_element_by_xpath(keyword_path).text.split("\n")[0]
                key_word=driver.find_element_by_xpath(keyword_path).text.split("\n")[1]
            if len(arr_pro)<int(key_num)+1:
                arr_pro.append([key_word])
            else:
                arr_pro[int(key_num)].append(key_word)
            # keyword_list.append(key_word)
        # 다음 페이지 넘기기
        driver.find_element_by_xpath('//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/a[2]').click()
    # return keyword_list

def pre_login(path):

    driver_execute(path)
    driver.find_element_by_xpath("//input[@id='loginId']").clear()
    driver.find_element_by_xpath("//input[@id='loginId']").send_keys("signcody4657")
    driver.find_element_by_xpath("//input[@id='loginPassword']").clear()
    driver.find_element_by_xpath("//input[@id='loginPassword']").send_keys("dsp3134233")
    driver.find_element_by_xpath("//input[@id='loginPassword']").send_keys(Keys.RETURN)

def pre_ad_login(path):

    driver_execute(path)
    driver.find_element_by_xpath("//a[@class='btn btn-lg btn-block btn-primary']").click()
    driver.find_element_by_xpath("//input[@id='uid']").clear()
    driver.find_element_by_xpath("//input[@id='uid']").send_keys("signcody")
    driver.find_element_by_xpath("//input[@id='upw']").clear()
    driver.find_element_by_xpath("//input[@id='upw']").send_keys("dsp3134233")
    driver.find_element_by_xpath("//input[@id='upw']").send_keys(Keys.RETURN)

def get_pr_report(path,arr_pro):

    driver_execute(path)
    element = driver.find_element_by_id("__naverpay")
    driver.switch_to.frame(element)
    time.sleep(1)
    driver.find_element_by_xpath("//a[@class='btn select_data']").click()
    driver.find_element_by_xpath('//div[@class="fix_range"]/ul/li[2]').click()
    driver.find_element_by_xpath('//span[@class="select_range"]').click()
    tmp_date1=""
    while True:
        tmp_dval=driver.find_element_by_xpath("//a[@class='btn select_data']/span/span[2]/em").get_attribute("innerHTML").replace(".","").replace(" ","")[2:]
        if tmp_date1!="" and tmp_date1==tmp_dval:
            break
        tmp_date1=tmp_dval
        tmp_date=tmp_dval[:2] + "/" + tmp_dval[2:4] + "월" + str((int(tmp_dval[4:6])+7)//7) + "주차"
        tmp_add=driver.find_element_by_xpath("//table[@class='tbl_list']/tbody/tr/td[5]").get_attribute("innerHTML")
        tmp_amt=driver.find_element_by_xpath("//table[@class='tbl_list']/tbody/tr/td[11]").get_attribute("innerHTML")
        tmp_buy=driver.find_element_by_xpath("//table[@class='tbl_list']/tbody/tr/td[13]").get_attribute("innerHTML")
        arr_pro.insert(0,[tmp_date,tmp_add*1,tmp_amt*1,tmp_buy*1])

        driver.find_element_by_xpath("//a[@class='btn pre']").click()

def get_pr_keyword(path):
    driver_execute(path)
    driver.find_element_by_xpath("//a[@data-nclicks-code='stu.selling']").click()
    driver.find_element_by_xpath('//div[@class="selectize-input items ng-valid ng-pristine has-options full has-items"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//div[@data-value="500"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//button[@class="btn btn-sm btn-default progress-button progress-button-dir-horizontal progress-button-style-top-line"]').click()
    time.sleep(1)

def get_ad_report(path):

    driver_execute(path)
    driver.find_element_by_xpath('//button[@class="btn-sm dropdown-toggle btn btn-default"]').click()
    driver.find_element_by_xpath("(//button[text()='필터 만들기'])").click()
    driver.find_element_by_xpath('//button[@style="word-break: keep-all;"]').click()
    driver.find_element_by_xpath('//div[@class="pl-2 filter-checkbox"][2]').click()
    driver.find_element_by_xpath('//button[@class="btn btn-sm btn-default-blue apply-button"]').click()
    driver.find_element_by_xpath("(//span[text()='다운로드'])").click()
    time.sleep(1)

def get_ad_mass(path):

    driver_execute(path)
    driver.find_element_by_xpath('//button[@class="btn btn-default dropdown-toggle"]').click()
    driver.find_element_by_xpath("(//button[text()='쇼핑검색'])").click()
    driver.find_element_by_xpath("(//button[@id='campaigns-dropdown'])").click()
    driver.find_element_by_xpath("(//div[@style='position: absolute; top: 0px; left: 0px; will-change: transform; transform: translate(0px, 32px);']/div[text()=' 스토어팜 ']/input)").click()
    driver.find_element_by_xpath("(//button[text()='다운로드'])").click()
    tmp=driver.find_element_by_xpath("(//tbody[@class='has-data ng-star-inserted']/tr[1]/td[11]/elena-mass-result)").get_attribute("innerHTML")
    while tmp=="<!---->":
        driver.find_element_by_xpath('//button[@class="btn btn-sm btn-default btn-icon"]').click()
        time.sleep(5)
        tmp=driver.find_element_by_xpath("(//tbody[@class='has-data ng-star-inserted']/tr[1]/td[11]/elena-mass-result)").get_attribute("innerHTML")
    driver.find_element_by_xpath("(//tbody[@class='has-data ng-star-inserted']/tr[1]/td[11]/elena-mass-result)").click()
    time.sleep(1)

def get_data_file():
    today_graph = nowDate + "_광고대비매출(주단위).csv"
    if find_file(today_graph)=="False":
        arr_pro=[]
        get_pr_report("https://sell.smartstore.naver.com/#/bizadvisor/marketing",arr_pro)
        df = pd.DataFrame(arr_pro)
        df.columns = ['주간', '광고비', '매출건수', '매출액']
        df.to_csv(FILE_FOLDER+"/"+today_graph, encoding='utf-8-sig')

    tmp_file=find_file("Product_")
    if tmp_file=="False" and not nowDate in tmp_file:
        get_pr_keyword("https://sell.smartstore.naver.com/#/products/origin-list")
        tmp_file=find_file("Product_")
        chg_file_arr=tmp_file.split("_")
        os.rename(FILE_FOLDER+"/"+tmp_file, FILE_FOLDER+"/"+chg_file_arr[1]+"_"+chg_file_arr[0]+"_"+chg_file_arr[2])

    arr_pro=[]
    arr_cate=[]
    tmp_file=find_file(nowDate + "_Product")
    if find_file(tmp_file)!="False":
        tmp_file=os.path.join(FILE_FOLDER,tmp_file)
        df=pd.read_csv(tmp_file, encoding='utf-8-sig')[["상품명","상품번호(스마트스토어)","할인가(PC)","대분류","중분류","소분류","세분류","대표이미지 URL"]]
        df["카테고리명"]=""
        df.loc[(df['세분류']==""),"카테고리명"]=df['대분류']+">>"+df['중분류']+">>"+df['소분류']
        df.loc[(df['세분류']!=""),"카테고리명"]=df['대분류']+">>"+df['중분류']+">>"+df['소분류']+">>"+df['세분류']
        df2=df[["상품명","상품번호(스마트스토어)"]].copy()
        df2["추가홍보문구1"]=""
        df2["추가홍보문구2"]=""
        df2["제외키워드"]=""
        tmp_file2=find_file("※상품별 부가정보")
        if find_file(tmp_file2)!="False":
            tmp_file2=os.path.join(FILE_FOLDER,tmp_file2)
            df2=pd.read_csv(tmp_file2, encoding='utf-8-sig')
        df=pd.merge(df,df2,how="left")
        df.to_csv(FILE_FOLDER + "/" + nowDate + "_상품 리스트_판매중.csv", encoding='utf-8-sig' ,index = False)
        df2.to_csv(FILE_FOLDER + "/※상품별 부가정보.csv", encoding='utf-8-sig' ,index = False)

    today_cate = "☆카테고리별 인기검색어.csv"
    if find_file(today_cate)=="False":
        arr_pro=[]
        arr_cate=arr_cate[1:]
        arr_pro.append(arr_cate)
        time.sleep(1)
        for cate_it in arr_cate:
            get_cate_key(cate_it,arr_pro)
        df = pd.DataFrame(arr_pro)
        columnNames = df.iloc[0]
        df = df[1:]
        df.columns = columnNames
        df.to_csv(FILE_FOLDER+'/'+startDate+'~'+nowDate+'_'+today_cate, encoding='utf-8-sig' ,index = False)

    zp_fname=find_file("mas")
    if zp_fname=="False":
    # 검색어
        get_ad_report("https://manage.searchad.naver.com/customers/392590/reports/rtt-a001-000000000451507")
    # 스토어팜
        get_ad_report("https://manage.searchad.naver.com/customers/392590/reports/rtt-a001-000000000451508")
    # 광고소재리스트
        get_ad_mass("https://manage.searchad.naver.com/customers/392590/tool/mass")
        zp_fname=find_file("mas")
        os.chdir(FILE_FOLDER)
        zipfile.ZipFile(zp_fname).extractall()
        os.remove(zp_fname)

        na_list=os.listdir(FILE_FOLDER)
        for na in na_list:
            if not nowDate in na and na != "import":
                os.rename(FILE_FOLDER+"/"+na, FILE_FOLDER+"/"+nowDate+"_광고시스템-"+na.replace("+","-"))

def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def make_chart(title,minR,maxR,cP1,cP2,wid):
    cats = Reference(ws, min_col=2, max_col=2, min_row=minR, max_row=maxR)
    value_ad = Reference(ws, min_col=3, max_col=3, min_row=minR, max_row=maxR)
    value_cnt = Reference(ws, min_col=4, max_col=4, min_row=minR, max_row=maxR)
    value_price = Reference(ws, min_col=5, max_col=5, min_row=minR, max_row=maxR)

    chart_ad = LineChart()
    chart_ad.add_data(value_ad)
    chart_ad.set_categories(cats)
    chart_ad.y_axis.axId = 200
    chart_ad.y_axis.title = '광고비'
    chart_ad.y_axis.majorGridlines = None
    s = chart_ad.series[0]
    s.graphicalProperties.line.solidFill = "0070C0"

    chart_cnt = LineChart()
    chart_cnt.add_data(value_cnt)
    chart_cnt.y_axis.title = "매출건"

    chart_cnt.title = title+"_매출건"
    chart_cnt.width = wid
    chart_cnt.height = 6
    chart_cnt.y_axis.crosses = "max"
    chart_cnt.legend.tagname="b"
    s = chart_cnt.series[0]
    s.graphicalProperties.line.solidFill = "F4B084"
    chart_cnt += chart_ad
    chart_cnt.legend=None

    chart_price = LineChart()
    chart_price.add_data(value_price)
    chart_price.y_axis.title = "매출액"

    chart_price.title = title+"_매출액"
    chart_price.width = wid+1
    chart_price.height = 6
    chart_price.y_axis.crosses = "max"
    s = chart_price.series[0]
    s.graphicalProperties.line.solidFill = "FFC000"
    chart_price += chart_ad
    chart_price.legend=None

    ws.add_chart(chart_cnt, cP1)
    ws.add_chart(chart_price, cP2)

def set_data_to_report():
# 엑셀데이터 가져오기
    global wb
    global ws

    wb = load_workbook(THIS_FOLDER+'\마케팅보고서.xlsx')
    ws = wb['보고서']
    for row in ws['L7:AC22']:
        for cell in row:
            cell.value = None

    # 유입키워드 컷트라인
    in_keyword_per_cut=ws['L5'].value
    in_keyword_cnt_cut=ws['L6'].value

    # 인기키워드 top 10
    arr_key_rank=[]
    arr_key_cnt=[]
    for row in ws['K13:K22']:
        for cell in row:
            tmp=int(cell.value.replace("~",""))
            if not tmp in arr_key_rank:
                arr_key_rank.append(tmp)
                arr_key_cnt.append(1)
            else:
                arr_key_cnt[len(arr_key_cnt)-1]=arr_key_cnt[len(arr_key_cnt)-1]+1
    # 제외키워드
    arr_key_exc=[]
    for i in range(25,27):
        cnt=11
        tmp_v=ws.cell(i,cnt).value
        while tmp_v!=None:
            arr_key_exc.append(tmp_v)
            cnt=cnt+1
            tmp_v=ws.cell(i,cnt).value
    arr_exc=("|").join(arr_key_exc)

# 유입 csv 데이터 추출
    tmp_file=find_file("검색어-보고서")
    tmp_file=os.path.join(FILE_FOLDER,tmp_file)
    df=pd.read_csv(tmp_file, encoding='utf-8-sig', skiprows=1)
    df=df[1:][["검색어","노출수","클릭수","전환수"]]
    columnNames = df.columns
    for i in range(1,4):
        col=columnNames[i]
        df[col] = df[col].astype(str).str.replace(',', '')
        df=df.astype({col: int})
        df=df.sort_values(by=col, ascending=False)
        df2=df[df[col] > df[col].quantile((100-in_keyword_per_cut)/100)]
        df2=df2.head(in_keyword_cnt_cut)
        ws.cell(6+i,11).value=col
        for j in range(0,df2.value_counts().size):
            ws.cell(6+i,12+j).value=(df2.iloc[j,0]+"\r\n"+str(df2.iloc[j,i]))
            ws.cell(6+i,12+j).alignment = Alignment(wrapText=True)

# 인기 csv 데이터 추출
    tmp_file=find_file("카테고리")
    tmp_file=os.path.join(FILE_FOLDER,tmp_file)
    df=pd.read_csv(tmp_file, encoding='utf-8-sig')
    for i in range(1,len(df.columns)):
        col=df.columns[i]
        tmp_col=col.split(">>")
        ws.cell(11,11+i).value=("\r\n").join(tmp_col[0:len(tmp_col)-1])
        ws.cell(11,11+i).alignment = Alignment(wrapText=True)
        ws.cell(12,11+i).value=tmp_col[len(tmp_col)-1]
        ws.column_dimensions[get_column_letter(11+i)].width = 10
    for x in range(1,len(df.columns)):
        pre=0
        pre_cnt=0
        for i in range(0,len(arr_key_rank)):
            if i>0:
                pre=arr_key_rank[i-1]
                pre_cnt=pre_cnt+arr_key_cnt[i-1]
            ws.column_dimensions[get_column_letter(12+x)].width = 10
            df2=df.iloc[pre:arr_key_rank[i],[0,x]]
            df2=df2.sample(arr_key_cnt[i])
            while (df2.iloc[:,1].str.contains(arr_exc)).any():
                df2=df.iloc[pre:arr_key_rank[i],[0,x]]
                df2=df2.sample(arr_key_cnt[i])
            df2=df2.sort_index()
            for y in range(0,df2.value_counts().size):
                mer_t=str(df2.iloc[y,1])+" > "+str(df2.iloc[y,0])
                ws.cell(13+y+pre_cnt,11+x).value=mer_t

# 광고대비 매출 csv 추출
    tmp_file=find_file("광고대비매출")
    tmp_file=os.path.join(FILE_FOLDER,tmp_file)
    df=pd.read_csv(tmp_file, encoding='utf-8-sig')

    for i in range(1,len(df.columns)):
        col=df.columns[i]
        if i==2 or i==4:
            df[col] = df[col].str.replace(',', '')
            df=df.astype({col: float})
    for x in range(1,len(df.columns)):
        for y in range(0,df.value_counts().size):
            ws.cell(6+y,1+x).value=df.iloc[y,x]
    make_chart("주간",df.value_counts().size+2,df.value_counts().size+5,"B6","F6",8)
    make_chart("전체",6,df.value_counts().size+5,"J28","J39",35)

    tail_df_a=[96,48,16]
    detail_df_a=[["반기","B36","F36",8],["분기","B25","F25",8],["월간","B14","F14",8]]
    idx=0
    last_y=y+1
    for tl in tail_df_a:
        df=df.tail(tl)
        del_tail=tail_df_a[idx]/4
        arr_tmp=[]
        for y in range(0,df.value_counts().size):
            # print(df.iloc[y])
            tmp_idx=(int(y//del_tail))
            if len(arr_tmp) < tmp_idx+1:
                arr_tmp.append([df.iloc[y,1],df.iloc[y,2],df.iloc[y,3],df.iloc[y,4]])
            else:
                for j in range(1,4):
                    arr_tmp[tmp_idx][j]=arr_tmp[tmp_idx][j]+df.iloc[y,j+1]
        df_tmp=df
        ws.cell(7+last_y,2).value=detail_df_a[idx][0]+" 부분합"
        for x in range(0,len(arr_tmp[0])):
            for y in range(0,len(arr_tmp)):
                ws.cell(8+last_y+y,2+x).value=arr_tmp[y][x]
        make_chart(detail_df_a[idx][0],8+last_y,8+last_y+y,detail_df_a[idx][1],detail_df_a[idx][2],detail_df_a[idx][3])
        last_y=last_y+y+3
        idx=idx+1

    max_ad=df.tail(4).iloc[:,2].max()
    next_month=nowDate[4:6]
    ws.cell(1,2).value=str(int(next_month)+1)+"월"
    ws.cell(2,3).value=int(max_ad)*4

    wb.save(FILE_FOLDER+'/'+startDate+'~'+nowDate+'_☆마케팅보고서.xlsx')

def check_warngg(p_id,g_id,s_id,s_name):
    path="https://manage.searchad.naver.com/customers/"+str(c_id)+"/adgroups/"+str(g_id)
    # print(driver.current_url)
    driver_execute(path)

    rslt_flg="False"
    if driver.find_element_by_xpath('(//button[@class="btn-sm btn-toggle dropdown-toggle btn btn-default"])').text!="행 표시: 200":
        driver.find_element_by_xpath('//button[@class="btn-sm btn-toggle dropdown-toggle btn btn-default"]').click()
        driver.find_element_by_xpath("(//button[text()=200])").click()
    # 행표시 200개 / 다음페이지 이동 / 노출가능 클릭
    row_posi = driver.find_element_by_xpath('//td[@data-value="'+s_id+'"]/preceding-sibling::td[1]/span/a')
    row_text=row_posi.text

    if row_text=="소재 연동제한":
        action = ActionChains(driver)
        action.move_to_element(driver.find_element_by_xpath('//div[@class="inner-left"]')).perform()
        driver.find_element_by_xpath('//td[@data-value="'+s_id+'"]/preceding-sibling::td[1]/preceding-sibling::td[1]/preceding-sibling::td[1]/input').click()
        action.move_to_element(driver.find_element_by_xpath('//div[@class="inner-left"]')).perform()
        driver.find_element_by_xpath('//button[@class="ml-2 btn btn-default btn-sm"]').click()
        driver.find_element_by_xpath('//button[@class="btn btn-primary Confirm_modal-button__1Kwk6 btn btn-secondary"]').click()
    elif row_text=="노출가능":
        if stillNew:
            warn_arr=[p_id,g_id,s_id,s_name,wait_val,"-"]
        else:
            warn_arr=[p_id,g_id,s_id,s_name,badgg_val,"-"]
        rslt_flg=pd.DataFrame([warn_arr],columns=['상품 ID','광고그룹 ID','소재 ID','노출상품명','제한상태','제한사유'])
    else:
        action = ActionChains(driver)
        try:
            action.move_to_element(row_posi).perform()
            row_posi.click()
        except Exception as e:
            row_posi_parent=driver.find_element_by_xpath('//td[@data-value="'+s_id+'"]/parent::tr/preceding-sibling::tr[1]/td[1]')
            action.move_to_element(row_posi_parent).perform()
            row_posi.click()
        row_detail=[]
        for row_detail_each in driver.find_elements_by_xpath('//div[@class="list-dot"]'):
            row_detail.append(row_detail_each.text.replace("더 알아보기",""))
        row_detail_img = driver.find_element_by_xpath('//td[@data-value="'+s_id+'"]/following-sibling::td/div/div/div/div[@class="image-preview"]').get_attribute("style")
        img_link=row_detail_img.split('"')[1]
        tmp_img_link=img_link.split(".")
        urllib.request.urlretrieve(img_link, OUTPUT_FOLDER+"\\"+s_id+"."+tmp_img_link[len(tmp_img_link)-1])
        warn_arr=[p_id,g_id,s_id,s_name,row_text,"\r\n".join(row_detail)]
        rslt_flg=pd.DataFrame([warn_arr],columns=['상품 ID','광고그룹 ID','소재 ID','노출상품명','제한상태','제한사유'])
        driver.find_element_by_xpath('//button[@class="modal-button btn btn-default"]').click()
    return rslt_flg

def set_data_to_ad():
# 엑셀데이터 가져오기
    global wb
    global ws
    global ws3
    global c_id
    global wait_val
    global badgg_val

    wb = load_workbook(THIS_FOLDER+'\광고관리.xlsx')
    ws = wb['검색광고']
    ws3 = wb['추가홍보문구']

    # 노출 상태 종류
    goodgg_val=ws['O2'].value
    goodgg_col=ws['O2'].fill.start_color.index
    badgg_val=ws['P2'].value
    badgg_col=ws['P2'].fill.start_color.index
    chapri_val=ws['Q2'].value
    chapri_col=ws['Q2'].fill.start_color.index
    wait_val=ws['R2'].value
    wait_col="FFC65911"
    newgg_val=ws['O3'].value
    newgg_col=ws['O3'].fill.start_color.index
    offgg_val=ws['P3'].value
    # offgg_col=ws['P3'].fill.start_color.index
    offgg_col="FFAEAAAA"
    warngg_val=ws['Q3'].value
    warngg_col=ws['Q3'].fill.start_color.index
    emptygg_val=ws['R3'].value
    # emptygg_col=ws['R3'].fill.start_color.index
    emptygg_col="FFFFFFFF"

    gg_val=[goodgg_val,badgg_val,chapri_val,wait_val,newgg_val,offgg_val,warngg_val,emptygg_val]
    gg_col=[goodgg_col,badgg_col,chapri_col,wait_col,newgg_col,offgg_col,warngg_col,emptygg_col]

    # 양호소재 컷트라인
    cut_rank=ws['C3'].value
    cut_show=ws['D3'].value
    cut_clk=ws['E3'].value
    cut_buy=ws['F3'].value
    cut_price_u=ws['G3'].value
    cut_price_d=ws['H3'].value

    # 입찰가 조절라인
    price_dwn_cut=ws['C5'].value
    price_dwn_val=ws['C6'].value
    price_up_cut=ws['D5'].value
    price_up_val=ws['D6'].value

    price_default=ws['F5'].value

# 상품 csv 데이터 추출
    pro_file=find_file("상품 리스트")
    pro_file=os.path.join(FILE_FOLDER,pro_file)
    df_pro=pd.read_csv(pro_file, encoding='utf-8-sig')
    df_pro.rename(columns = {'상품번호(스마트스토어)' : '상품 ID', '할인가(PC)' : '판매가'}, inplace = True)

    mas_file=find_file("mas")
    mas_file=os.path.join(FILE_FOLDER,mas_file)
    df_mas=pd.read_csv(mas_file, encoding='utf-8-sig', skiprows=1)
    df_mas.rename(columns = {'쇼핑몰 상품ID' : '상품 ID'}, inplace = True)
    df_mas['노출상품명']=df_mas['노출상품명'].fillna(df_mas['기본상품명'])
    c_id=df_mas["CUST_ID"].iat[0]

    report_file=find_file("스토어팜-보고서")
    report_file=os.path.join(FILE_FOLDER,report_file)
    df_report=pd.read_csv(report_file, encoding='utf-8-sig', skiprows=1)
    df_report.rename(columns = {'소재' : '소재 ID', '총비용(VAT포함,원)': '광고비용', '전환매출액(원)': '전환액', '전환율(%)': '전환율' ,'광고그룹': '광고그룹 이름'}, inplace = True)
    sum_col_val=['노출수','클릭수','광고비용','전환수', '전환율', '전환액']
    report_col_val=['노출상태',   '노출상품명',  '광고그룹 이름',  '평균노출순위', '노출수',  '클릭수',  '전환수', '전환율',  '광고비용', '전환액',     '소재 상태',    '소재 입찰가',   '소재 ID',  '광고그룹 ID', '상품 ID']
    report_col_val_del=['\r\n',  '(',           ')\r\n',         '/',           '/',      '/',       '/',      '/',      '/',        '\r\n',      '/',            "\r\n",         '/',         '/',           ""]
    total_col_val=['제외키워드','상품명','상품 ID','대표이미지 URL','판매가','대분류','중분류','소분류','세분류','노출수','클릭수','광고비용','전환수','전환율','전환액']

    x=10
    for sum_val in sum_col_val:
        df_report[sum_val] = df_report[sum_val].astype(str).str.replace(',', '')
        df_report=df_report.astype({sum_val: float})
        if sum_val=="전환율":
            tmp_sum=(df_report["전환수"].sum()*10000/df_report["클릭수"].sum()//1)/100
            ws.cell(10,x).value=int(tmp_sum)
        else:
            tmp_sum=df_report[sum_val].sum()
            ws.cell(10,x).value=int(tmp_sum)
        x=x+1

    df_mas=pd.merge(df_mas,df_report,on=["소재 ID","광고그룹 이름"],how="outer")
    df_mas=df_mas.sort_values(by=['소재 상태','상품 ID','클릭수','광고비용','노출수','광고그룹 이름'], ascending=[False,True,False,True,False,False])

    df_mas_sum=df_mas.groupby('상품 ID').sum().reset_index()[['상품 ID','노출수','클릭수','광고비용','전환수','전환액','전환율']].copy()
    df_mas_sum=df_mas_sum.sort_values(by=['노출수','클릭수','광고비용'], ascending=[False,False,True])
    df_mas_sum["전환율"]=(df_mas_sum["전환수"]*10000/df_mas_sum["클릭수"]//1)/100

    df_grp_arr=df_mas.groupby('광고그룹 이름').sum().reset_index().copy()['광고그룹 이름'].values

    df_mas_each=df_mas.copy()
    # df_mas_each=df_mas[df_mas["상품 ID"].notna()].copy()
    df_mas_each["노출상태"]=goodgg_val
    df_mas_each=df_mas_each[report_col_val]
    df_mas_each.loc[(df_mas_each['평균노출순위']>=price_dwn_cut),"노출상태"]=chapri_val+"인상 " + df_mas_each["소재 입찰가"].astype(str) + ">" + (df_mas_each["소재 입찰가"]+price_dwn_val).astype(str) + ""
    df_mas_each.loc[(df_mas_each['평균노출순위']<=price_up_cut),"노출상태"]=chapri_val+"인하 " + df_mas_each["소재 입찰가"].astype(str) + ">" + (df_mas_each["소재 입찰가"]+price_up_val).astype(str) + ""
    df_mas_each.loc[(df_mas_each['평균노출순위']<=cut_rank) & (df_mas_each['노출수']<=cut_show) & (df_mas_each['클릭수']<=cut_clk) & (df_mas_each['전환수']<=cut_buy) & ~((cut_price_d <= df_mas_each['소재 입찰가']) & (df_mas_each['소재 입찰가'] <= cut_price_u)),"노출상태"]=badgg_val
    df_mas_each.loc[(pd.isnull(df_mas_each['노출수'])),"노출상태"]=warngg_val

    df_pid_arr=df_mas_each['상품 ID'].copy().drop_duplicates().values
    for df_pid in df_pid_arr:
        tmp_df=df_mas_each[df_mas_each['상품 ID']==df_pid]
        df_grp_arr_each=tmp_df['광고그룹 이름'].copy().drop_duplicates().values
        grp_minus=np.setdiff1d(df_grp_arr,df_grp_arr_each)
        for x in grp_minus:
            next_each=len(df_mas_each)
            df_mas_each.loc[next_each,"노출상태"]=emptygg_val
            df_mas_each.loc[next_each,"광고그룹 이름"]=x
            df_mas_each.loc[next_each,"광고그룹 ID"]=df_mas[df_mas['광고그룹 이름']==x]['광고그룹 ID'].iloc[0]
            df_mas_each.loc[next_each,"상품 ID"]=df_pid
    df_mas_each=df_mas_each.sort_values(by=['상품 ID','노출수','클릭수','광고비용'], ascending=[False,False,False,True])
    df_mas_each=df_mas_each.reset_index(drop=True)

    df_pid_arr=df_mas_each['상품 ID'].copy().drop_duplicates().values
    for df_pid in df_pid_arr:
        tmp_df=df_mas_each[df_mas_each['상품 ID']==df_pid].copy()
        tmp_df_idx=df_mas_each.index[df_mas_each['상품 ID']==df_pid].copy()
        cnt_bad=0
        for sub_idx in range(0,len(tmp_df)):
            each_idx=tmp_df_idx[sub_idx]
            if sub_idx<10:
                tmp_what_show=tmp_df.iloc[sub_idx]["노출상태"]
                if badgg_val in tmp_what_show:
                    cnt_bad=cnt_bad+1
                elif offgg_val in tmp_what_show or emptygg_val in tmp_what_show:
                    if not tmp_df['노출상태'].iat[0] in [badgg_val,warngg_val]:
                        df_mas_each['노출상태'].iat[each_idx]=newgg_val
                        df_mas_each['상품 ID'].iat[each_idx]=tmp_df.iloc[0]['상품 ID']
                        df_mas_each['소재 상태'].iat[each_idx]="on"
                        df_mas_each['소재 입찰가'].iat[each_idx]=price_default
            elif tmp_df['노출상태'].iat[sub_idx]!=emptygg_val:
                if cnt_bad==0:
                    df_mas_each['노출상태'].iat[each_idx]=offgg_val
                else:
                    df_mas_each['노출상태'].iat[each_idx]=newgg_val
                    cnt_bad=cnt_bad-1
    df_mas_each=df_mas_each.sort_values(by=['상품 ID','노출수','클릭수','광고비용'], ascending=[False,False,False,True])

    warn_each_file=find_file("☆점검소재리스트")
    if warn_each_file!="False":
        warn_each_file=os.path.join(FILE_FOLDER,warn_each_file)
        df_warn_detail=pd.read_csv(warn_each_file, encoding='utf-8-sig')
    else:
        df_warn_detail=pd.DataFrame(columns=['광고그룹 ID','소재 ID','노출상품명','제한상태','제한사유'])
    df_warn_sid_arr=df_warn_detail['소재 ID'].values
    for g_name in df_grp_arr:
        g_id=df_mas_each[df_mas_each['광고그룹 이름']==g_name]['광고그룹 ID'].iloc[0]
        df_sid=df_mas_each[(df_mas_each['노출상태']==warngg_val) & (df_mas_each['광고그룹 이름']==g_name) & (df_mas_each['소재 상태']=="on")].copy()
        df_sid_arr=df_sid['소재 ID'].values
        df_sname_arr=df_sid['노출상품명'].values
        s_idx=0
        for s_id in df_sid_arr:
            if not s_id in df_warn_sid_arr:
                s_name=df_sname_arr[s_idx]
                p_id=df_mas_each[df_mas_each['소재 ID']==s_id]['상품 ID'].iloc[0]
                tmp_df=check_warngg(p_id,g_id,s_id,s_name)
                if isinstance(tmp_df, pd.DataFrame):
                    df_warn_detail=df_warn_detail.append(tmp_df, ignore_index = True)
                s_idx=s_idx+1
    df_warn_detail=df_warn_detail.sort_values(by=['제한사유','상품 ID','제한상태'], ascending=[False,False,False])
    df_warn_detail.to_csv(OUTPUT_FOLDER+'/'+startDate+'~'+nowDate+'_☆점검소재리스트.csv' , encoding='utf-8-sig',index = False)

# todo 신규소재 키워드 부여
    df_rel_key=""
    rel_key_file=find_file("☆상품 연관키워드")
    if rel_key_file!="False":
        rel_key_file=os.path.join(FILE_FOLDER,rel_key_file)
        df_rel_key=pd.read_csv(rel_key_file, encoding='utf-8-sig')
    df_p_id=df_mas_each[df_mas_each['노출상태']==newgg_val]['상품 ID'].drop_duplicates().values
    df_show_name=df_mas_each[(df_mas_each['상품 ID'].isin(df_p_id))]['노출상품명'].drop_duplicates().values

    if not isinstance(df_rel_key, pd.DataFrame):
        df_rel_key=pd.DataFrame(columns=(call_RelKwd("test").columns.insert(0,"노출키워드")))
    df_rel_key_arr=df_rel_key['노출키워드'].drop_duplicates().values
    for each_show in df_show_name:
        tmp_each_show=str(each_show).split(" ")
        for each_show_2 in tmp_each_show:
            if not each_show_2 in df_rel_key_arr and not each_show_2 in [",", ".", "/", "(", ")", "", "-"]:
                df_rel_key_each=call_RelKwd(str(each_show_2))
                if isinstance(df_rel_key_each, pd.DataFrame):
                    df_rel_key_each['노출키워드']=each_show_2
                    df_rel_key=df_rel_key.append(df_rel_key_each, ignore_index = True)
                    df_rel_key=df_rel_key.drop_duplicates(subset=['노출키워드','연관키워드'])
# 적당히 자동화때 삭제
                    df_rel_key.to_csv(OUTPUT_FOLDER+'/'+startDate+'~'+nowDate+'_☆상품 연관키워드.csv', encoding='utf-8-sig' ,index = False)

    df_rel_key.to_csv(FILE_FOLDER+'/'+startDate+'~'+nowDate+'_☆상품 연관키워드.csv', encoding='utf-8-sig',index = False)
    df_rel_key.to_csv(OUTPUT_FOLDER+'/'+startDate+'~'+nowDate+'_☆상품 연관키워드.csv', encoding='utf-8-sig' ,index = False)

    df_pid_arr=df_mas_each['상품 ID'].copy().drop_duplicates().values
    for df_pid in df_pid_arr:
        tmp_df=df_mas_each[df_mas_each['상품 ID']==df_pid].copy()
        tmp_df_idx=df_mas_each.index[df_mas_each['상품 ID']==df_pid].copy()
        for sub_idx in range(0,len(tmp_df)):
            each_idx=tmp_df_idx[sub_idx]
            if newgg_val in tmp_df['노출상태'].iat[sub_idx]:
                prev_name = tmp_df['노출상품명'].iat[0]
                new_name=make_new_name(prev_name,df_rel_key)
                df_mas_each['노출상품명'].iat[each_idx]=new_name
            elif warngg_val in tmp_df['노출상태'].iat[sub_idx]:
                tmp_warn=df_warn_detail[df_warn_detail['소재 ID']==tmp_df['소재 ID'].iat[sub_idx]]
                if tmp_warn['제한상태'].iloc[0]!="소재 노출제한":
                    df_mas_each['노출상태'].iat[each_idx]=tmp_warn['제한상태'].iloc[0]

    df_mas_each.to_csv(FILE_FOLDER+"/"+nowDate+"_광고소재리스트.csv", encoding='utf-8-sig',index = False)
    df_mas_each.to_csv(OUTPUT_FOLDER+"/"+nowDate+"_광고소재리스트.csv", encoding='utf-8-sig',index = False)

    # todo 상품등록
    df_product_add=df_mas_each[df_mas_each["노출상태"]==newgg_val][["광고그룹 ID","상품 ID","소재 입찰가","노출상품명"]].copy()
    df_product_add.to_csv(OUTPUT_FOLDER+"/"+nowDate+"_신규소재등록.csv", encoding='utf-8-sig',index = False)
    # todo 입찰가 변경
    df_product_chgpri=df_mas_each[df_mas_each["노출상태"].str.contains(chapri_val)].copy()
    df_product_chgpri["소재 입찰가"]=df_product_chgpri["노출상태"].str.split(">").str[1]
    df_product_chgpri=df_product_chgpri[["광고그룹 ID","소재 ID","소재 입찰가"]]
    df_product_chgpri.to_csv(OUTPUT_FOLDER+"/"+nowDate+"_입찰가변경.csv", encoding='utf-8-sig',index = False)
    # todo on/off 소재 일괄 변경
    df_product_2on = df_mas_each[df_mas_each["노출상태"]==offgg_val][["광고그룹 ID","소재 ID","소재 상태"]].copy()
    df_product_2on["소재 상태"]="off"
    df_product_2off= df_mas_each[df_mas_each["노출상태"]!=offgg_val][["광고그룹 ID","소재 ID","소재 상태"]].copy()
    df_product_2off["소재 상태"]="on"
    df_product_2on.append(df_product_2off)
    df_product_2on.to_csv(OUTPUT_FOLDER+"/"+nowDate+"_활성상태변경.csv", encoding='utf-8-sig',index = False)

    for df_idx in range(0,len(df_mas_sum)):
        this_pro_id=df_mas_sum.iloc[df_idx]['상품 ID']
        tmp_df=df_mas_each[df_mas_each['상품 ID'].astype(str)==str(this_pro_id.astype(float))].copy()
        this_pro_gid=df_mas_each[df_mas_each['상품 ID'].astype(str)==str(this_pro_id)].copy()
        for sub_idx in range(0,len(tmp_df)):
            if not "소재"+str(sub_idx+1) in df_mas_sum.columns:
                total_col_val.append("소재"+str(sub_idx+1))
                df_mas_sum["소재"+str(sub_idx+1)]=""
            tmp_df_val=""
            v_idx=0
            for each_col in report_col_val:
                tmp_df_val = tmp_df_val + str(tmp_df.iloc[sub_idx][each_col]).replace(".0","")
                tmp_df_val = tmp_df_val + report_col_val_del[v_idx]
                v_idx=v_idx+1
            df_mas_sum["소재"+str(sub_idx+1)].iat[df_idx]=str(tmp_df_val)
    df_all=pd.merge(df_pro,df_mas_sum,on="상품 ID",how="outer")[total_col_val]
    df_all=df_all.sort_values(by=['노출수','클릭수','광고비용'], ascending=[False,False,True])
    df_all['상품명']=df_all['상품명'].fillna("삭제상품")
    df_all.to_csv(FILE_FOLDER+"/"+nowDate+"_광고소재부분합리스트.csv", encoding='utf-8-sig',index = False)

    rate=0.9
    tmp_df=df_all[(df_all['노출수'] >= df_all['노출수'].quantile(rate)) & (df_all['전환수'] <= df_all['전환수'].quantile(rate))]
    tmp_df=tmp_df.sort_values(by="클릭수", ascending=True)
    tmp_df2=df_warn_detail.drop_duplicates("상품 ID")
    tmp_df=pd.merge(tmp_df,tmp_df2,how="left")
    tmp_df=tmp_df.head(4)
    tmp_df.to_csv(FILE_FOLDER+"/"+nowDate+"연습.csv", encoding='utf-8-sig',index = False)

    y=12
    for df_y in range(0,len(df_all)):
        x=2
        for df_x in range(0,len(df_all.iloc[df_y])):
            ws.cell(y,x).value=df_all.iloc[df_y,df_x]
            if "소재" in total_col_val[df_x]:
                ws.column_dimensions[colnum_string(x)].width = 30
                ws.row_dimensions[y].height = 48
                ws.cell(y,x).alignment = Alignment(wrapText=True)
                delim=str(ws.cell(y,x).value).split("\r\n")[0]
                for gg in gg_val:
                    gg_idx=gg_val.index(gg)
                    if gg in delim:
                        ws.cell(y,x).fill=PatternFill("solid", fgColor=gg_col[gg_idx])
            x=x+1
        y=y+1

    x=2
    for total_val in total_col_val:
        ws.cell(11,x).value=total_val
        if "소재1" == total_val:
            v_idx=0
            tmp_df_val=""
            for each_col in report_col_val:
                tmp_df_val = tmp_df_val + each_col
                tmp_df_val = tmp_df_val + report_col_val_del[v_idx]
                v_idx=v_idx+1
                ws.cell(8,x).value=tmp_df_val+"\r\n(노출▶클릭순▶)"
                ws.cell(8,x).alignment = Alignment(wrapText=True)
                ws.row_dimensions[8].height = 115
        x=x+1

    wb.save(FILE_FOLDER+'/'+startDate+'~'+nowDate+'_☆광고관리.xlsx')

def init():
    global THIS_FOLDER
    global FILE_FOLDER
    global OUTPUT_FOLDER
    global nowDate
    global startDate
    global stillNew
    global driver_exist

    driver_exist=False
    now = datetime.datetime.now()
    nowDate = now.strftime('%Y%m%d')
    startDate = nowDate
    stillNew = True
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    FILE_FOLDER = THIS_FOLDER+'\자료'
    createFolder(FILE_FOLDER)
    filedel_list=os.listdir(FILE_FOLDER)
    day_check=7
    for filedel in filedel_list:
        if not "※" in filedel:
            if "☆" in filedel:
                startD = datetime.datetime.strptime(filedel.split("_")[0].split("~")[0], "%Y%m%d")
                date_diff = now - startD
                if date_diff.days>day_check-2:
                    stillNew = False
                if date_diff.days>day_check:
                    del_fname=os.path.join(FILE_FOLDER,filedel)
                    os.remove(del_fname)
                    del_foldname=os.path.join(FILE_FOLDER,"import")
                    shutil.rmtree(del_foldname)
                elif not "import" in filedel:
                    before_fname=os.path.join(FILE_FOLDER,filedel)
                    after_fname=startD.strftime('%Y%m%d')+"~"+nowDate+"_"+filedel.split("_")[1]
                    os.rename(before_fname, FILE_FOLDER+"/"+after_fname)
                startDate = startD.strftime('%Y%m%d')
            elif not nowDate in filedel and not "import" in filedel:
                del_fname=os.path.join(FILE_FOLDER,filedel)
                os.remove(del_fname)
    OUTPUT_FOLDER = FILE_FOLDER+'\import'
    createFolder(OUTPUT_FOLDER)

def driver_execute(path):
    global driver_exist
    if not driver_exist:
        driver_init()
        driver_exist=True
        time.sleep(1)
        pre_login("https://sell.smartstore.naver.com/#/bizadvisor/marketing")
        time.sleep(1)
        pre_ad_login("https://manage.searchad.naver.com/customers/392590/reports/rtt-a001-000000000451507")
    if driver.current_url!=path:
        time.sleep(1)
        driver.get(path)

def driver_init():
    global driver

    g_name=os.path.join(THIS_FOLDER,"chromedriver")
    options = webdriver.ChromeOptions()
    # 창 숨기는 옵션 추가
    # options.add_argument("headless")
    prefs = {
      "download.default_directory": FILE_FOLDER,
      "download.prompt_for_download": False,
      "download.directory_upgrade": True,
      "safebrowsing.enabled": True
    }
    options.add_experimental_option('prefs', prefs)

    driver = webdriver.Chrome(g_name, options=options)
    driver.implicitly_wait(20)

def timer_start():
    global start_t
    start_t = time.time()

def timer_chk():
    print("time :", time.time() - start_t)

timer_start()
init()

get_data_file()
set_data_to_report()
set_data_to_ad()

if driver_exist:
    driver.close()
timer_chk()
